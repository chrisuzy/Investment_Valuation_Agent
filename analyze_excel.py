"""Analyze the ground truth Excel file: layout, input vs calculated cells, colors."""

import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(
    "knowledge_base/groud_truth/ground_truth.xlsx", data_only=False
)


def rgb_hex(color):
    if color is None:
        return None
    if hasattr(color, "rgb") and color.rgb and str(color.rgb) != "00000000":
        return str(color.rgb)
    if hasattr(color, "theme") and color.theme is not None:
        return f"theme:{color.theme}"
    if hasattr(color, "indexed") and color.indexed is not None:
        return f"idx:{color.indexed}"
    return None


def dump_sheet(ws, name):
    print(f"\n====== {name} ======")
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            val = cell.value
            if val is None:
                continue

            fill = cell.fill
            fg = rgb_hex(fill.fgColor) if fill and fill.fgColor else None
            font_color = rgb_hex(cell.font.color) if cell.font and cell.font.color else None
            font_bold = cell.font.bold if cell.font else False

            is_formula = str(val).startswith("=") if val else False

            col_letter = get_column_letter(c)
            cell_ref = f"{col_letter}{r}"

            val_str = str(val)[:100]
            tags = []
            if fg:
                tags.append(f"fill:{fg}")
            if font_color:
                tags.append(f"font:{font_color}")
            if font_bold:
                tags.append("BOLD")
            if is_formula:
                tags.append("FORMULA")

            tag_str = f"  [{', '.join(tags)}]" if tags else ""
            print(f"  {cell_ref}: {val_str}{tag_str}")


# Dump the key sheets
for sheet_name in [
    "Input sheet",
    "Valuation output",
    "R& D converter",
    "Operating lease converter",
    "Cost of capital worksheet",
    "Option value",
    "Diagnostics",
]:
    dump_sheet(wb[sheet_name], sheet_name)
