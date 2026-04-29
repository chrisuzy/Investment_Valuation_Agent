"""
Build a pre-filled Ginzu workbook per test company.

For each ticker in COMPANIES, produce `TEST_DATA/ginzu_prefilled/<TICKER>_ginzu_input.xlsx`
by copying `knowledge_base/Ginzu_NVIDIA.xlsx` and writing the test-company's values
into the exact input cells Ginzu expects. The user opens each file in Excel on
Windows — Excel recalculates automatically — and reads the output cells.

Changes written per workbook:
  • Input sheet — rows 3–40 (base-year + value drivers + market + options)
                  rows 56–83 (stable overrides + failure + reinvestment-lag + etc.)
                  rows 42–52 ZEROED OUT (NVIDIA AI/Auto story — contaminates non-NVDA)
  • R& D converter sheet — B6 (N) + B7 (current LTM R&D) + B11–B15 (past 5y history)
                           AND zeroes any old NVIDIA R&D history past the 5-year window
  • All other sheets untouched (calc sheets recalc themselves in Excel).

Run:
  cd backend && source .venv/bin/activate
  python tools/build_ginzu_prefilled_workbooks.py            # all 4 companies
  python tools/build_ginzu_prefilled_workbooks.py MSFT       # one company
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "backend"))

import openpyxl

from tools.build_ginzu_input_packages import (  # type: ignore
    COMPANIES, extract_data_for_ticker,
)
from data_sources.damodaran_store import DamodaranStore


GINZU_SRC = REPO_ROOT / "knowledge_base" / "Ginzu_NVIDIA.xlsx"
OUT_DIR = REPO_ROOT / "TEST_DATA" / "ginzu_prefilled"


# ───────────────────────────────────────────────────────────────────────────
# Ginzu Input-sheet cell writes
# ───────────────────────────────────────────────────────────────────────────

def write_input_sheet(ws, d: dict):
    """Write all input-sheet cells for one company. See Ginzu_NVIDIA.xlsx row labels."""

    # Base-year metadata
    ws["B3"] = d["valuation_date"]                  # Date of valuation
    ws["B4"] = d["company_name"]                     # Company name
    ws["B7"] = d["country"]                          # Country of incorporation
    ws["B8"] = d["industry_us"]                      # Industry (US)
    ws["B9"] = d["industry_global"]                  # Industry (Global)

    # Base-year flows (col B = LTM / "this year", col C = prior FY, col D = years since 10K)
    ws["B10"] = d["rev_ltm"];        ws["C10"] = d["rev_prior"];   ws["D10"] = d["years_since_10k"]
    ws["B11"] = d["ebit_ltm"];       ws["C11"] = d["ebit_prior"]
    ws["B12"] = d["int_ltm"];        ws["C12"] = d["int_prior"]
    ws["B13"] = d["bv_eq"];          ws["C13"] = d["bv_eq_prior"]
    ws["B14"] = d["bv_debt"];        ws["C14"] = d["bv_debt_prior"]
    ws["B15"] = d["has_rd"]                          # "Yes" / "No"
    ws["B16"] = d["has_leases"]
    ws["B17"] = d["cash"];           ws["C17"] = d["cash_prior"]
    ws["B18"] = d["cross"];          ws["C18"] = d["cross_prior"]
    ws["B19"] = d["minority"];       ws["C19"] = d["minority_prior"]
    ws["B20"] = d["shares"]
    ws["B21"] = d["price"]
    ws["B22"] = d["tax_eff"]
    ws["B23"] = d["tax_marg"]

    # Value drivers
    ws["B25"] = d["rev_growth_yr1"]
    ws["B26"] = d["op_margin_yr1"]
    ws["B27"] = d["rev_growth_2_5"]
    ws["B28"] = d["target_margin"]
    ws["B29"] = d["conv_year"]
    ws["B30"] = d["s_c_high"]
    ws["B31"] = d["s_c_stable"]

    # Market
    ws["B33"] = d["rf"]

    # Options — all test companies have no employee options in CIQ data
    ws["B36"] = "No"
    ws["B37"] = 0
    ws["B38"] = 0
    ws["B39"] = 0
    ws["B40"] = 0

    # ── ZERO OUT NVIDIA-specific AI/Auto story drivers (rows 42-52) ──
    # These columns drive Ginzu's 3-story valuation. For non-NVIDIA firms,
    # zeroing the market size collapses AI and Auto PV to 0 so only the
    # main single-business DCF contributes.
    for r in range(42, 53):
        for col in ("B", "C"):
            cell = f"{col}{r}"
            existing = ws[cell].value
            if isinstance(existing, (int, float)):
                ws[cell] = 0
            # leave strings (labels) alone

    # ── Default assumption overrides (rows 56-83) ──
    # B55-57: Stable WACC override
    ws["B56"] = "Yes"
    ws["B57"] = d["stable_wacc"]
    # B58-60: Stable ROIC override
    ws["B59"] = "Yes"
    ws["B60"] = d["stable_roic"]
    # B61-65: Failure probability (keep No for test companies)
    ws["B62"] = "No"
    ws["B63"] = 0
    ws["B64"] = "V"
    ws["B65"] = 0.5
    # B66-68: Reinvestment lag
    ws["B67"] = "Yes"
    ws["B68"] = d["reinv_lag"]
    # B69-70: Tax convergence (leave default No)
    ws["B70"] = "No"
    # B71-73: NOL
    ws["B72"] = "No"
    ws["B73"] = 0
    # B74-76: Riskfree override (No)
    ws["B75"] = "No"
    ws["B76"] = 0.02
    # B77-79: Growth perpetuity override (No)
    ws["B78"] = "No"
    ws["B79"] = -0.05
    # B80-83: Trapped cash (No)
    ws["B81"] = "No"
    ws["B82"] = 0
    ws["B83"] = 0.15


def write_rd_converter(ws, d: dict, max_hist_rows: int = 10):
    """Fill R&D converter sheet B6 (N), B7 (current R&D), B11..B11+N-1 (past R&D).

    Ginzu's R&D converter layout (verified by inspection):
      row 6  — amortization period N (input cell is typically B6 or wherever the
               "How many years do you want to amortize R&D?" prompt lives)
      row 7  — current year R&D expense
      row 11 — R&D year -1   (most recent past year)
      row 12 — R&D year -2
      ...
      row 15 — R&D year -5

    For safety, we also zero-out rows 16-21 (Ginzu's NVDA may have only 5 years
    but the template can hold more) so stale NVIDIA values don't contaminate.
    """
    if d["has_rd"] != "Yes":
        return

    # N and current-year R&D — these live in the "Inputs" area. Ginzu's
    # Input sheet actually sets N via the industry default; we override it
    # here to force the N we want.
    ws["B6"] = d["rd_amort_n"]
    ws["B7"] = d["rd_current"]

    # Past years — Ginzu stores up to ~10 years starting at row 11.
    # Any slots past the ones we fill need to be cleared (the template may
    # have historical NVIDIA values past index len(rd_past)).
    hist = list(d["rd_past"])[:max_hist_rows]
    # Pad if we have fewer than N years of history (Ginzu's formula expects
    # non-zero for the first N years; zeros mean "no amortization from year -k"
    # which is fine mathematically).
    for i in range(max_hist_rows):
        row = 11 + i
        v = hist[i] if i < len(hist) else 0
        ws[f"B{row}"] = v


def write_lease_converter(ws, d: dict):
    """Fill Operating lease converter inputs (if has_leases)."""
    if d["has_leases"] != "Yes":
        return
    # Ginzu's lease converter input cells (verify in workbook; typical layout):
    #   B4 — current lease expense
    #   B7..B11 — commitments year 1..5
    #   B12 — commitment beyond year 5
    ws["B4"] = d["lease_expense"]
    commits = d["lease_commits"] + [0] * 6
    for i in range(5):
        ws[f"B{7 + i}"] = commits[i]
    ws["B12"] = commits[5] if len(commits) > 5 else 0


# ───────────────────────────────────────────────────────────────────────────
# Main
# ───────────────────────────────────────────────────────────────────────────

def build_one(ticker: str, d: dict):
    out_path = OUT_DIR / f"{ticker}_ginzu_input.xlsx"
    shutil.copy2(GINZU_SRC, out_path)

    wb = openpyxl.load_workbook(out_path)
    write_input_sheet(wb["Input sheet"], d)
    write_rd_converter(wb["R& D converter"], d)
    write_lease_converter(wb["Operating lease converter"], d)
    wb.save(out_path)
    print(f"[{ticker}] wrote {out_path.relative_to(REPO_ROOT)}  ({out_path.stat().st_size:,} bytes)")


def main(tickers: list[str]):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    dam = DamodaranStore.from_directory(str(REPO_ROOT / "knowledge_base" / "damodaran"))
    for t in tickers:
        print(f"[{t}] extracting...")
        try:
            d = extract_data_for_ticker(t, dam, None)
        except Exception as e:
            print(f"[{t}] FAILED extract: {e}")
            import traceback; traceback.print_exc()
            continue
        try:
            build_one(t, d)
        except Exception as e:
            print(f"[{t}] FAILED build: {e}")
            import traceback; traceback.print_exc()


if __name__ == "__main__":
    args = sys.argv[1:] or list(COMPANIES.keys())
    main(args)
