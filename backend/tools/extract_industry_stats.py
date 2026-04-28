"""Extract Input Stat Distributioons from Ginzu_NVIDIA.xlsx → JSON.

The Ginzu workbook ships with a sheet of industry-level quartile distributions
for revenue growth, operating margin, sales-to-capital, cost of capital, beta,
and debt-to-capital. These are the Ginzu I/J/K/L/M/N columns Damodaran uses
as benchmark references alongside user-hypothesis inputs.

Output schema (backend/data_sources/industry_stats.json):
{
  "industries": {
    "<industry_name>": {
      "n_firms": int,
      "revenue_growth_3y":      {"q1": float, "median": float, "q3": float},
      "pretax_operating_margin":{"q1": float, "median": float, "q3": float},
      "sales_to_capital":       {"q1": float, "median": float, "q3": float},
      "cost_of_capital":        {"q1": float, "median": float, "q3": float},
      "beta_median": float,
      "debt_to_capital":        {"q1": float, "median": float, "q3": float}
    },
    ...
  },
  "source": "Ginzu_NVIDIA.xlsx::Input Stat Distributioons",
  "extracted_at": "<iso>"
}
"""
from __future__ import annotations

import datetime as _dt
import json
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = REPO_ROOT / "knowledge_base" / "Ginzu_NVIDIA.xlsx"
OUTPUT_JSON = REPO_ROOT / "backend" / "data_sources" / "industry_stats.json"


def _num(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main() -> None:
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb["Input Stat Distributioons"]

    industries: dict[str, dict] = {}
    # Rows start at 3 (row 1 = headers, row 2 = sub-headers, rows 3+ = data)
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value  # col A
        if not name or not isinstance(name, str):
            continue
        n_firms = _num(ws.cell(row=r, column=2).value)
        if n_firms is None:
            continue

        # Columns per ginzu_extracted.json inspection:
        #   C/D/E: Revenue Growth 3Y — Q1 / Median / Q3
        #   F/G/H: Pre-tax Op Margin — Q1 / Median / Q3
        #   I/J/K: Sales to Invested Capital — Q1 / Median / Q3
        #   L/M/N: Cost of Capital — Q1 / Median / Q3
        #   O: Beta median
        #   P/Q/R: Debt to Capital Ratio — Q1 / Median / Q3
        def _q(col_q1: int) -> dict[str, float | None]:
            return {
                "q1": _num(ws.cell(row=r, column=col_q1).value),
                "median": _num(ws.cell(row=r, column=col_q1 + 1).value),
                "q3": _num(ws.cell(row=r, column=col_q1 + 2).value),
            }

        industries[name.strip()] = {
            "n_firms": int(n_firms),
            "revenue_growth_3y": _q(3),       # C
            "pretax_operating_margin": _q(6), # F
            "sales_to_capital": _q(9),        # I
            "cost_of_capital": _q(12),        # L
            "beta_median": _num(ws.cell(row=r, column=15).value),  # O
            "debt_to_capital": _q(16),        # P
        }

    out = {
        "industries": industries,
        "source": "Ginzu_NVIDIA.xlsx::Input Stat Distributioons",
        "extracted_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "industry_count": len(industries),
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_JSON.open("w") as f:
        json.dump(out, f, indent=1)

    print(f"Wrote {OUTPUT_JSON.relative_to(REPO_ROOT)}: {len(industries)} industries")


if __name__ == "__main__":
    main()
