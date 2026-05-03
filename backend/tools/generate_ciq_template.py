"""
Prepare a CIQ template for download by copying the user's authoritative
template and substituting the requested ticker.

Authoritative template: ``knowledge_base/ciq_fetches/CIQ_Fetch_Template_new.xlsx``
— formulas, mnemonics, layout, and currency handling are all authored there.
This script does NOT regenerate any formulas. It only:

  1. Copies the authoritative file to ``CIQ_Fetch_Template.xlsx``.
  2. Writes the ticker into cell ``B1`` of the ``CIQ_Data`` sheet (the
     $B$1-referencing formulas resolve off this cell).
  3. Swaps the legacy hardcoded ticker inside the geo-segment formulas
     (the ``IQ_GEO_SEG_*`` rows embed the ticker as a string literal
     rather than ``$B$1``, so a text substitution is required to make
     them resolve for a different company).

Usage:
  python -m tools.generate_ciq_template           # defaults to NVDA
  python -m tools.generate_ciq_template NVDA
  python -m tools.generate_ciq_template SEHK:992
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl


# ── Paths ──────────────────────────────────────────────────────
CIQ_DIR = Path(__file__).resolve().parent.parent.parent / "knowledge_base" / "ciq_fetches"
SOURCE_TEMPLATE = CIQ_DIR / "CIQ_Fetch_Template_new.xlsx"
OUTPUT_TEMPLATE = CIQ_DIR / "CIQ_Fetch_Template.xlsx"

# The authoritative template was authored with Lenovo HK ("992") as the
# example ticker. Geo-segment formulas embed this literal rather than
# referencing $B$1, so we need to substitute it on download.
LEGACY_HARDCODED_TICKER = "992"


def _read_legacy_ticker(ws) -> str:
    """Return whatever ticker is currently sitting in B1 of the source file.

    Falls back to ``LEGACY_HARDCODED_TICKER`` if B1 is missing or empty.
    Using whatever is actually in B1 keeps us robust to the user updating
    the source template with a different example ticker later.
    """
    b1 = ws["B1"].value
    if b1 is None:
        return LEGACY_HARDCODED_TICKER
    return str(b1).strip()


def generate_template(ticker: str = "NVDA") -> Path:
    """Prepare a download-ready CIQ template for ``ticker``.

    The source file's formulas are preserved exactly; only the ticker
    literal (B1 plus any legacy hardcoded occurrences) is changed.
    """
    if not SOURCE_TEMPLATE.exists():
        raise FileNotFoundError(
            f"Authoritative CIQ template not found: {SOURCE_TEMPLATE}. "
            f"This file is required — do not regenerate it from scratch."
        )

    # 1. Byte-copy the authoritative file so the output keeps all of
    #    Excel's XLL-binding metadata, the _CIQHiddenCacheSheet, and any
    #    other plugin-authored bits that openpyxl might not preserve
    #    perfectly on a cold save.
    shutil.copy2(SOURCE_TEMPLATE, OUTPUT_TEMPLATE)

    # 2. Open the copy and perform only the ticker substitution.
    wb = openpyxl.load_workbook(OUTPUT_TEMPLATE, data_only=False)
    ws = wb["CIQ_Data"]

    legacy_ticker = _read_legacy_ticker(ws)
    ws["B1"] = ticker

    # 3. Replace the legacy hardcoded ticker wherever it's embedded as a
    #    quoted string literal inside a formula. This should only hit the
    #    ~20 geo-segment cells that don't use $B$1.
    needle_quoted = f'"{legacy_ticker}"'
    replacement_quoted = f'"{ticker}"'
    replacements = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            if not v.startswith("="):
                continue
            if needle_quoted in v:
                cell.value = v.replace(needle_quoted, replacement_quoted)
                replacements += 1

    wb.save(OUTPUT_TEMPLATE)
    print(f"Prepared CIQ template: {OUTPUT_TEMPLATE}")
    print(f"  Ticker set to: {ticker} (was: {legacy_ticker})")
    print(f"  Hardcoded-ticker substitutions: {replacements}")
    print()
    print("Next steps:")
    print(f"  1. Open {OUTPUT_TEMPLATE.name} in Excel with the CIQ plugin")
    print(f"  2. Wait for all formulas to resolve")
    print(f"  3. Save the file")
    print(f"  4. Upload via the app's 'Load from CIQ File' button")

    return OUTPUT_TEMPLATE


if __name__ == "__main__":
    ticker = sys.argv[1] if len(sys.argv) > 1 else "NVDA"
    generate_template(ticker)
