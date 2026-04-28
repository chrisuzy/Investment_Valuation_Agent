"""Analyze all Damodaran Excel files and produce _catalog.json.

For each file: reads sheet names, headers, row count.
Also categorizes by dataset and region.

Usage:
    python backend/tools/analyze_damodaran.py
"""
import json
import re
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent.parent
DAM_DIR = BASE_DIR / "knowledge_base" / "damodaran"

# Known regions and their suffixes
REGION_SUFFIXES = {
    "": "US",
    "Global": "Global",
    "China": "China",
    "India": "India",
    "Japan": "Japan",
    "Europe": "Europe",
    "emerg": "Emerging",
    "Rest": "Rest",
}


def extract_dataset_region(filename: str) -> tuple[str, str]:
    """Extract base dataset name and region from filename."""
    stem = Path(filename).stem
    for suffix, region in sorted(REGION_SUFFIXES.items(), key=lambda x: -len(x[0])):
        if suffix and stem.endswith(suffix):
            return stem[: -len(suffix)], region
    return stem, "US"


def analyze_file(filepath: Path) -> dict:
    """Analyze a single Excel file."""
    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            headers = []
            row_count = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c else "" for c in row]
                row_count += 1
            sheets.append({
                "name": name,
                "headers": headers[:20],  # first 20 columns
                "row_count": row_count,
            })
        wb.close()
        dataset, region = extract_dataset_region(filepath.name)
        return {
            "filename": filepath.name,
            "dataset": dataset,
            "region": region,
            "sheets": sheets,
        }
    except Exception as e:
        return {
            "filename": filepath.name,
            "error": str(e),
        }


def main():
    files = sorted(DAM_DIR.glob("*.xls")) + sorted(DAM_DIR.glob("*.xlsx"))
    # Exclude catalog itself
    files = [f for f in files if f.name != "_catalog.json"]

    catalog = {}
    datasets = {}
    for i, f in enumerate(files, 1):
        info = analyze_file(f)
        catalog[f.name] = info
        ds = info.get("dataset", f.stem)
        if ds not in datasets:
            datasets[ds] = []
        datasets[ds].append(info.get("region", "unknown"))
        if i % 50 == 0:
            print(f"  Analyzed {i}/{len(files)}...")

    # Write catalog
    out_path = DAM_DIR / "_catalog.json"
    with open(out_path, "w") as fh:
        json.dump(catalog, fh, indent=2)
    print(f"\nCatalog written to {out_path}")
    print(f"Total files: {len(catalog)}")
    print(f"Unique datasets: {len(datasets)}")
    print("\nDatasets and regions:")
    for ds, regions in sorted(datasets.items()):
        print(f"  {ds}: {', '.join(sorted(regions))}")


if __name__ == "__main__":
    main()
