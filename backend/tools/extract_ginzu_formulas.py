"""Extract every cell from Ginzu_NVIDIA.xlsx into a label-enriched JSON corpus.

Output schema (docs/brainstorm_cache/ginzu_extracted.json):

{
  "metadata": {
    "source_file": "<path>",
    "extracted_at": "<iso>",
    "sheet_count": int,
    "formula_count": int,
    "value_count": int
  },
  "sheets": {
    "<sheet_name>": {
      "dimensions": "<A1:Z999>",
      "labels_by_row": { "<row_str>": "<first-non-empty-text-from-col-A-or-B>" },
      "cells": [
        {
          "cell": "B12",
          "row": 12,
          "col": "B",
          "value": <raw formula or literal>,
          "cached_value": <data-only cached result>,
          "is_formula": bool,
          "row_label": "<label from labels_by_row>",
          "row_label_source": "A"|"B"|null
        },
        ...
      ]
    },
    ...
  }
}

Goal: give downstream walkers a single JSON they can grep by variable label
instead of by cell address.
"""
from __future__ import annotations

import datetime as _dt
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = REPO_ROOT / "knowledge_base" / "Ginzu_NVIDIA.xlsx"
OUTPUT_JSON = REPO_ROOT / "docs" / "brainstorm_cache" / "ginzu_extracted.json"


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _build_row_labels(ws_formula) -> tuple[dict[str, str], dict[str, str]]:
    """For each row that has any content, record the first non-empty text cell
    in column A (preferred) or column B (fallback) as the row label.
    Returns (labels_by_row, source_by_row)."""
    labels: dict[str, str] = {}
    sources: dict[str, str] = {}
    max_row = ws_formula.max_row or 0
    for r in range(1, max_row + 1):
        for col_letter in ("A", "B"):
            cell = ws_formula[f"{col_letter}{r}"]
            val = cell.value
            if isinstance(val, str) and val.strip() and not _is_formula(val):
                labels[str(r)] = val.strip()
                sources[str(r)] = col_letter
                break
    return labels, sources


def _extract_sheet(ws_formula, ws_data) -> dict:
    labels_by_row, source_by_row = _build_row_labels(ws_formula)

    cells: list[dict] = []
    max_row = ws_formula.max_row or 0
    max_col = ws_formula.max_column or 0

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            formula_cell = ws_formula.cell(row=r, column=c)
            raw = formula_cell.value
            if raw is None or (isinstance(raw, str) and not raw.strip()):
                continue

            data_cell = ws_data.cell(row=r, column=c)
            cached = data_cell.value

            col_letter = get_column_letter(c)
            addr = f"{col_letter}{r}"
            is_f = _is_formula(raw)

            # Don't duplicate the pure-label cells as "cells" entries — they
            # are already captured in labels_by_row. Skip column A and column B
            # cells that are strings (labels) and NOT formulas — but keep them
            # if they look like a value rather than a label (e.g. numbers or
            # formulas in A/B). This keeps the list focused on data + compute.
            if (
                col_letter in ("A", "B")
                and isinstance(raw, str)
                and not is_f
            ):
                continue

            entry = {
                "cell": addr,
                "row": r,
                "col": col_letter,
                "value": raw if not isinstance(raw, _dt.datetime) else raw.isoformat(),
                "cached_value": cached if not isinstance(cached, _dt.datetime) else cached.isoformat(),
                "is_formula": is_f,
                "row_label": labels_by_row.get(str(r)),
                "row_label_source": source_by_row.get(str(r)),
            }
            cells.append(entry)

    return {
        "dimensions": ws_formula.dimensions,
        "labels_by_row": labels_by_row,
        "cells": cells,
    }


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_XLSX}")

    wb_formula = load_workbook(SOURCE_XLSX, data_only=False)
    wb_data = load_workbook(SOURCE_XLSX, data_only=True)

    sheets_out: dict[str, dict] = {}
    formula_count = 0
    value_count = 0

    for name in wb_formula.sheetnames:
        ws_f = wb_formula[name]
        ws_d = wb_data[name]
        sheet_info = _extract_sheet(ws_f, ws_d)
        sheets_out[name] = sheet_info
        for cell in sheet_info["cells"]:
            if cell["is_formula"]:
                formula_count += 1
            else:
                value_count += 1

    out = {
        "metadata": {
            "source_file": str(SOURCE_XLSX.relative_to(REPO_ROOT)),
            "extracted_at": _dt.datetime.now().isoformat(timespec="seconds"),
            "sheet_count": len(sheets_out),
            "formula_count": formula_count,
            "value_count": value_count,
        },
        "sheets": sheets_out,
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_JSON.open("w") as f:
        json.dump(out, f, indent=1, default=str)

    print(
        f"Wrote {OUTPUT_JSON.relative_to(REPO_ROOT)}: "
        f"{len(sheets_out)} sheets, "
        f"{formula_count} formulas, "
        f"{value_count} values."
    )


if __name__ == "__main__":
    main()
