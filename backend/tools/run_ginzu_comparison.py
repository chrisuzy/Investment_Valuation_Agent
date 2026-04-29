"""
Phase 3 — compare our backend vs Ginzu ground truth for one test company.

Workflow:
  1. Parse `docs/experiments/ginzu_inputs_<TICKER>.md` — read Section A cells
     AND the Section D table where the user has filled in Ginzu's computed values.
  2. Build `CompanyValuationInput` with identical assumptions (Section A values).
  3. POST to /api/valuation → get our backend output.
  4. Produce `docs/experiments/ginzu_comparison_<TICKER>.md` with per-module tables:
     ours | ginzu | Δ absolute | Δ % | flag.
  5. Update `docs/experiments/ginzu_comparison_summary.md` with a cross-company top-3.

Run:
  cd backend && source .venv/bin/activate
  python tools/run_ginzu_comparison.py MSFT
  python tools/run_ginzu_comparison.py              # all companies with filled-in packages
"""

from __future__ import annotations

import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "backend"))

# Reuse the package extractor for Section A values (keeps them in lockstep)
from tools.build_ginzu_input_packages import (  # type: ignore
    COMPANIES, DEFAULT_ASSUMPTIONS, extract_data_for_ticker, OUTPUT_CELLS_SPEC
)
from data_sources.damodaran_store import DamodaranStore


BACKEND = "http://localhost:8000"
MATCH_PCT = 0.01   # ✓  — within 1%
SMALL_PCT = 0.05   # ⚠  — within 5%
                   # ❌  — ≥ 5%


def _read_ginzu_values_from_excel(xlsx_path: Path) -> dict[tuple[str, str], float | None]:
    """Read Ginzu's computed output values directly from the recalculated workbook.

    Expects `<TICKER>_ginzu_output.xlsx` — the pre-filled workbook after the user
    opened it in Excel, let it recalc, and saved. `data_only=True` reads the
    cached result values (not the formulas).

    Returns {(sheet, cell): value} for every (sheet, cell) in OUTPUT_CELLS_SPEC.
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    out: dict[tuple[str, str], float | None] = {}
    for sheet, cell, _label, _module in OUTPUT_CELLS_SPEC:
        if sheet not in wb.sheetnames:
            out[(sheet, cell)] = None
            continue
        v = wb[sheet][cell].value
        if isinstance(v, (int, float)):
            out[(sheet, cell)] = float(v)
        else:
            out[(sheet, cell)] = None
    return out


def _build_backend_input(ticker: str, d: dict[str, Any]) -> dict:
    """Construct the /api/valuation body using THE SAME values from the package so the
    comparison is apples-to-apples (identical assumptions)."""
    return {
      "ticker":             ticker,
      "company_name":       d["company_name"],
      "country":            d["country"],
      "reporting_currency": "USD",
      "stock_price_currency": "USD",
      "quarters_since_10k": 0,   # package collapses to LTM=FY0
      "period_date_10k":    d["valuation_date"],
      "period_date_10q":    d["valuation_date"],
      "period_dates_annual": {"0": d["valuation_date"]},
      "effective_tax_rate_ciq": d["tax_eff"],
      "raw_financials": [
        {
          "fiscal_year": 2025, "revenues": d["rev_ltm"], "ebit": d["ebit_ltm"],
          "ebitda": None, "net_income": None,
          "interest_expense": d["int_ltm"], "capex": None, "d_a": None,
          "noncash_wc": None, "change_in_noncash_wc": None, "net_debt_issued": None,
          "cash_and_marketable_securities": d["cash"],
          "bv_equity": d["bv_eq"], "bv_debt": d["bv_debt"],
          "mv_equity": d["shares"] * d["price"], "mv_debt": None,
          "shares_outstanding": d["shares"], "stock_price": d["price"],
          "cross_holdings": d["cross"], "minority_interests": d["minority"],
          "r_and_d_expense": d["rd_current"],
          "earnings_before_tax": None, "total_tax_expense": None,
        },
        {
          "fiscal_year": 2024, "revenues": d["rev_prior"], "ebit": d["ebit_prior"],
          "ebitda": None, "net_income": None, "interest_expense": d["int_prior"],
          "capex": None, "d_a": None, "noncash_wc": None,
          "change_in_noncash_wc": None, "net_debt_issued": None,
          "cash_and_marketable_securities": d["cash_prior"],
          "bv_equity": d["bv_eq_prior"], "bv_debt": d["bv_debt_prior"],
          "mv_equity": None, "mv_debt": None,
          "shares_outstanding": d["shares"],  # assume constant
          "stock_price": None,
          "cross_holdings": d["cross_prior"], "minority_interests": d["minority_prior"],
          "r_and_d_expense": d["rd_past"][0] if d["rd_past"] else 0,
          "earnings_before_tax": None, "total_tax_expense": None,
        },
      ],
      "quarterly_financials": [],
      "adjustment_inputs": {
        "amortization_period_n": d["rd_amort_n"],
        "r_and_d_expense_current": d["rd_current"],
        "r_and_d_expense_past": d["rd_past"],
        "operating_lease_expense_current": d["lease_expense"],
        "operating_lease_commitments": d["lease_commits"],
        "has_r_and_d":      d["has_rd"] == "Yes",
        "has_operating_leases": d["has_leases"] == "Yes",
      },
      "macro_inputs": {
        "risk_free_rate":    d["rf"],
        "equity_risk_premium": d["erp"],
        "country_risk_premium": d["crp"],
        "tax_rate_marginal": d["tax_marg"],
        "tax_rate_effective": d["tax_eff"],
        "default_spread":    0.0085,
      },
      "industry_data": {
        "industry_name":     d["industry_us"],
        "region": "US",
        "beta_u": 1.0,   # filled by backend industry lookup anyway; unused here
        "beta_u_corrected_for_cash": None,
        "industry_d_e_ratio": None, "industry_effective_tax_rate": None,
        "cost_of_equity": None, "cost_of_debt_pretax": 0.055, "wacc": None,
        "pretax_operating_margin": None, "after_tax_operating_margin": None,
        "sales_to_capital": d["s_c_high"], "revenue_growth": None,
        "std_dev_stock": None, "roic": None,
        "ev_ebitda": None, "ev_sales": None, "pe_ratio": None, "pbv_ratio": None,
      },
      "industry_data_global": None,
      "company_metrics": None,
      "option_inputs": {
        "number_of_options": 0, "average_strike_price": 0,
        "average_maturity": 0, "stock_price_std_dev": 0,
        "dividend_yield": 0, "has_options": False,
      },
      "valuation_assumptions": {
        "projection_years": 10, "high_growth_years": 5,
        "stable_growth_rate": None,
        "revenue_growth_next_year":   d["rev_growth_yr1"],
        "operating_margin_next_year": d["op_margin_yr1"],
        "revenue_growth_years_2_5":   d["rev_growth_2_5"],
        "target_operating_margin":    d["target_margin"],
        "margin_convergence_year":    d["conv_year"],
        "sales_to_capital_high":      d["s_c_high"],
        "sales_to_capital_stable":    d["s_c_stable"],
        "cost_of_capital_stable_override": d["stable_wacc"],
        "roic_stable_override":       d["stable_roic"],
        "failure_probability": 0.0,
        "distress_proceeds_pct": 0.5, "failure_tie_to": "V",
        "override_reinvestment_lag": True,
        "reinvestment_lag_years": d["reinv_lag"],
        "override_tax_convergence": False, "override_nol": False, "nol_amount": 0,
        "override_riskfree": False, "riskfree_after_yr10": None,
        "override_growth_perpetuity": False, "growth_perpetuity_rate": None,
        "override_trapped_cash": False, "trapped_cash_amount": 0,
        "trapped_cash_tax_rate": 0,
      },
      "methodology_choices": {
        "cost_of_capital_approach": "detailed",
        "wacc_direct_input": None,
        "beta_approach": "single_business_us", "beta_direct_input": None,
        "erp_approach": "country_of_incorporation", "erp_direct_input": None,
        "kd_approach": "industry_fallback", "kd_direct_input": None,
        "synthetic_rating_firm_type": "large", "actual_rating": None,
        "decile_region": "US", "decile_risk_group": "Median",
        "debt_maturity_years": 5.0, "use_bond_pricing_for_debt": False,
        "business_segments": [], "geographic_segments": [],
        "convertible_debt": {"book_value":0,"interest_expense":0,"maturity_years":0,"market_value":0},
        "has_convertible": False,
        "preferred_stock": {"shares":0,"price_per_share":0,"dividend_per_share":0},
        "has_preferred": False,
        "unsupported_branch_warnings": [],
      },
    }


def _post_valuation(inputs: dict) -> dict:
    r = subprocess.run(
        ["curl", "-s", "-X", "POST", f"{BACKEND}/api/valuation",
         "-H", "Content-Type: application/json",
         "-d", json.dumps({"inputs": inputs})],
        capture_output=True, text=True, timeout=120,
    )
    return json.loads(r.stdout)


def _fetch_from_file(ticker: str) -> dict:
    """Use the real fetch-from-file flow so industry lookup + LTM + country
    macro go through production code paths."""
    path = REPO_ROOT / "TEST_DATA" / f"TEST_{ticker}.xlsx"
    r = subprocess.run(
        ["curl", "-s", "-X", "POST", f"{BACKEND}/api/valuation/fetch-from-file",
         "-F", f"file=@{path}", "-F", "region=US", "-F", "risk_free_rate=0.0425"],
        capture_output=True, text=True, timeout=180,
    )
    return json.loads(r.stdout)


def _patch(sid: str, overrides: dict) -> dict:
    r = subprocess.run(
        ["curl", "-s", "-X", "PATCH", f"{BACKEND}/api/valuation/{sid}",
         "-H", "Content-Type: application/json",
         "-d", json.dumps({"overrides": overrides})],
        capture_output=True, text=True, timeout=120,
    )
    return json.loads(r.stdout)


# Map each (sheet, cell) in OUTPUT_CELLS_SPEC → field in our backend response
# so we can look up "what does our backend say" for every Ginzu metric.
def _our_value(sheet: str, cell: str, resp: dict) -> float | None:
    cc = resp.get("cost_of_capital") or {}
    adj = resp.get("adjusted") or {}
    dcf = resp.get("dcf") or {}
    fin = resp.get("final") or {}

    key = f"{sheet}:{cell}"
    table = {
        # M2
        "Cost of capital worksheet:B23": cc.get("beta_u"),
        "Cost of capital worksheet:C57": cc.get("beta_l"),
        "Cost of capital worksheet:B27": cc.get("equity_risk_premium"),
        "Cost of capital worksheet:B37": cc.get("cost_of_debt_pretax"),
        "Cost of capital worksheet:B60": cc.get("mv_equity"),
        "Cost of capital worksheet:C60": cc.get("mv_debt_total"),
        "Cost of capital worksheet:B61": cc.get("weight_equity"),
        "Cost of capital worksheet:C61": cc.get("weight_debt"),
        "Cost of capital worksheet:B62": cc.get("cost_of_equity"),
        "Cost of capital worksheet:C62": cc.get("cost_of_debt_aftertax"),
        "Cost of capital worksheet:E62": cc.get("wacc"),
        "Cost of capital worksheet:B13": cc.get("wacc"),

        # M1 R&D
        "R& D converter:B24": adj.get("unamortized_r_and_d", 0) + adj.get("amortization_r_and_d", 0),  # approximation
        "R& D converter:D35": adj.get("value_of_research_asset"),
        "R& D converter:D37": adj.get("amortization_r_and_d"),
        "R& D converter:D39": None,  # EBIT adjustment — need to recompute from raw
        # M1 Leases
        "Operating lease converter:F31": adj.get("depreciation_on_lease_asset"),
        "Operating lease converter:F32": adj.get("lease_adjustment_to_ebit"),
        "Operating lease converter:F33": adj.get("pv_of_operating_leases"),

        # M4 — year-by-year
        "Summary Sheet:B3":  (dcf.get("revenue_projections") or [None])[0],
        "Summary Sheet:B12": (dcf.get("revenue_projections") or [None]*10)[9] if len(dcf.get("revenue_projections") or []) >= 10 else None,
        "Summary Sheet:E3":  (dcf.get("ebit_projections") or [None])[0],
        "Summary Sheet:E12": (dcf.get("ebit_projections") or [None]*10)[9] if len(dcf.get("ebit_projections") or []) >= 10 else None,
        "Summary Sheet:F16": (dcf.get("fcff_projections") or [None])[0],
        "Summary Sheet:F25": (dcf.get("fcff_projections") or [None]*10)[9] if len(dcf.get("fcff_projections") or []) >= 10 else None,
        "Summary Sheet:D3":  None,  # operating margin — ours is revenue × margin; extract if needed
        "Summary Sheet:D12": None,

        # M4 — terminal + going concern
        "Valuation output:C30": cc.get("wacc"),  # terminal cost of capital (our simple engine uses same as yr10)
        "Valuation output:C31": (dcf.get("discount_factors") or [None]*10)[9] if len(dcf.get("discount_factors") or []) >= 10 else None,
        "Valuation output:B40": dcf.get("value_of_operating_assets"),

        # M7 bridge
        "Valuation output:B43": dcf.get("value_of_operating_assets"),
        "Valuation output:B48": dcf.get("value_of_equity"),
        "Valuation output:B50": dcf.get("value_of_equity"),

        # M9 per-share
        "Valuation output:B52": fin.get("value_per_share"),

        # M8 options
        "Option value:B28": fin.get("call_value_per_option"),
        "Option value:B29": fin.get("value_of_all_options"),
    }
    return table.get(key)


def _flag(ours: float | None, theirs: float | None) -> str:
    if ours is None or theirs is None: return "—"
    if theirs == 0 and ours == 0:       return "✓"
    if theirs == 0:                      return "❌"
    pct = abs((ours - theirs) / theirs)
    if pct < MATCH_PCT: return "✓"
    if pct < SMALL_PCT: return "⚠"
    return "❌"


def _fmt(v: float | None, pct_mode=False) -> str:
    if v is None: return "—"
    if pct_mode:
        return f"{v*100:.2f}%"
    if abs(v) >= 1000 or abs(v) < 0.001:
        return f"{v:,.2f}"
    return f"{v:.4f}"


def compare(ticker: str) -> str:
    """Produce the comparison markdown for one ticker.

    Reads Ginzu's outputs directly from
      TEST_DATA/ginzu_prefilled/<TICKER>_ginzu_output.xlsx
    which is the recalculated version of the pre-filled input workbook.
    (User opens `<TICKER>_ginzu_input.xlsx`, saves as `_output.xlsx` after Excel recalcs.)

    Falls back to `<TICKER>_ginzu_input.xlsx` if `_output.xlsx` doesn't exist —
    Excel writes cached values back into the source file on save, so either name works.
    """
    prefill_dir = REPO_ROOT / "TEST_DATA" / "ginzu_prefilled"
    output_xlsx = prefill_dir / f"{ticker}_ginzu_output.xlsx"
    input_xlsx  = prefill_dir / f"{ticker}_ginzu_input.xlsx"
    xlsx_path = output_xlsx if output_xlsx.exists() else input_xlsx
    if not xlsx_path.exists():
        return f"# {ticker}\n\nNo Ginzu workbook found at `{xlsx_path}`.\n"

    ginzu_vals = _read_ginzu_values_from_excel(xlsx_path)
    n_filled = sum(1 for v in ginzu_vals.values() if v is not None)
    if n_filled == 0:
        return (f"# {ticker}\n\nGinzu workbook at `{xlsx_path}` has no cached "
                f"computed values — user must open it in Excel and save after "
                f"recalculation. Skipping.\n")

    # Use the REAL production flow: fetch-from-file goes through the industry
    # resolver + LTM rotation + country macro lookup. Then PATCH to align the
    # analyst's methodology choices (ERP approach, Kd approach, assumptions) to
    # match whatever the user typed into the Ginzu workbook.
    resp = _fetch_from_file(ticker)
    if "detail" in resp:
        return f"# {ticker}\n\nBackend error: {resp['detail']}\n"
    sid = resp["id"]

    # Pull key inputs straight from the USER'S GINZU WORKBOOK — these are the
    # single source of truth for what assumptions to push to our backend.
    import openpyxl
    gwb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    gin = gwb["Input sheet"]

    def _g(cell):
        v = gin[cell].value
        return float(v) if isinstance(v, (int, float)) else v

    # Build the override bundle using Ginzu's input-sheet values so comparison
    # is apples-to-apples on assumptions (not just methodology).
    overrides = {
        # Value drivers (Input sheet rows 25-31)
        "valuation_assumptions.revenue_growth_next_year": _g("B25"),
        "valuation_assumptions.operating_margin_next_year": _g("B26"),
        "valuation_assumptions.revenue_growth_years_2_5": _g("B27"),
        "valuation_assumptions.target_operating_margin": _g("B28"),
        "valuation_assumptions.margin_convergence_year": int(_g("B29") or 5),
        "valuation_assumptions.sales_to_capital_high": _g("B30"),
        "valuation_assumptions.sales_to_capital_stable": _g("B31"),
        "macro_inputs.risk_free_rate": _g("B33"),
        "macro_inputs.tax_rate_marginal": _g("B23"),
        "effective_tax_rate_ciq": _g("B22"),
        # Stable-period overrides (rows 56-60)
        "valuation_assumptions.cost_of_capital_stable_override": _g("B57") if _g("B56") == "Yes" else None,
        "valuation_assumptions.roic_stable_override":            _g("B60") if _g("B59") == "Yes" else None,
        # Reinvestment lag (rows 67-68)
        "valuation_assumptions.override_reinvestment_lag":       _g("B67") == "Yes",
        "valuation_assumptions.reinvestment_lag_years":          int(_g("B68") or 1),
    }

    # Methodology choices — read from Cost-of-capital worksheet
    coc = gwb["Cost of capital worksheet"]
    beta_approach = str(coc["B21"].value or "").strip()
    erp_approach  = str(coc["B25"].value or "").strip()
    kd_approach   = str(coc["B33"].value or "").strip()

    # Map Ginzu's labels to our enum values
    BETA_MAP = {"Single Business(US)": "single_business_us",
                "Single Business(Global)": "single_business_global",
                "Multibusiness(US)": "multi_business_us",
                "Multibusiness(Global)": "multi_business_global",
                "Direct input": "direct_levered"}
    ERP_MAP  = {"Country of incorporation": "country_of_incorporation",
                "Operating countries":      "operating_countries",
                "Operating regions":        "operating_regions",
                "Will input":               "direct"}
    KD_MAP   = {"Industry average": "industry_fallback",
                "Direct input":     "direct",
                "Synthetic rating": "synthetic_rating",
                "Actual rating":    "actual_rating"}

    if beta_approach in BETA_MAP:
        overrides["methodology_choices.beta_approach"] = BETA_MAP[beta_approach]
    if erp_approach in ERP_MAP:
        overrides["methodology_choices.erp_approach"] = ERP_MAP[erp_approach]
        # If Ginzu used operating-countries/regions, it computes a weighted ERP.
        # We can't reproduce the weighting without the underlying revenue mix,
        # so override with direct input = Ginzu's computed ERP from C27.
        ginzu_erp = coc["B27"].value
        if erp_approach in ("Operating countries", "Operating regions") and isinstance(ginzu_erp, (int, float)):
            overrides["methodology_choices.erp_approach"]    = "direct"
            overrides["methodology_choices.erp_direct_input"] = float(ginzu_erp)
    if kd_approach in KD_MAP:
        overrides["methodology_choices.kd_approach"] = KD_MAP[kd_approach]
        if kd_approach == "Actual rating":
            rating = coc["B35"].value
            if rating:
                overrides["methodology_choices.actual_rating"] = str(rating)

    # Clean None values (PATCH should leave those untouched)
    overrides = {k: v for k, v in overrides.items() if v is not None}

    resp = _patch(sid, overrides)
    if "detail" in resp:
        return f"# {ticker}\n\nBackend error on PATCH: {resp['detail']}\n"

    # Walk through OUTPUT_CELLS_SPEC in order, compare each
    company_name = (resp.get("inputs", {}).get("company_name") or ticker)
    lines = [f"# Ginzu vs Backend — {ticker} ({company_name})", ""]
    lines.append(f"**Source:** `TEST_DATA/TEST_{ticker}.xlsx`  ·  **Ginzu package:** `ginzu_inputs_{ticker}.md`")
    lines.append("")
    lines.append(f"Ginzu values filled in: **{n_filled} / {len(OUTPUT_CELLS_SPEC)}**")
    lines.append("")

    by_module: dict[str, list] = {}
    for sheet, cell, label, module in OUTPUT_CELLS_SPEC:
        by_module.setdefault(module, []).append((sheet, cell, label))

    tot_match = tot_small = tot_big = 0
    for module, rows in by_module.items():
        lines.append(f"## {module}")
        lines.append("")
        lines.append("| Metric | Sheet · Cell | Ours | Ginzu | Δ (abs) | Δ (%) | Flag |")
        lines.append("|--------|--------------|------|-------|---------|-------|------|")
        # Pct metrics: only rates/weights/ratios, NOT betas or $ values.
        pct_metrics = {
            "B27",  # ERP used
            "B37",  # Kd pre-tax
            "B61", "C61",  # weights
            "B62", "C62", "E62", "B13",  # Ke, Kd_at, WACC
            "B41",  # failure probability
            "B54",  # Price as % of value
            "D3", "D12",  # operating margins
            "C30", "C31",  # terminal WACC, cumulative discount factor (fraction)
        }
        for sheet, cell, label in rows:
            ours = _our_value(sheet, cell, resp)
            theirs = ginzu_vals.get((sheet, cell))
            pct = cell in pct_metrics
            delta_abs = (ours - theirs) if (ours is not None and theirs is not None) else None
            delta_pct_num = (delta_abs / theirs * 100) if (delta_abs is not None and theirs not in (None, 0)) else None
            delta_pct_str = f"{delta_pct_num:+.1f}%" if delta_pct_num is not None else "—"
            flag = _flag(ours, theirs)
            if flag == "✓":
                tot_match += 1
            elif flag == "⚠":
                tot_small += 1
            elif flag == "❌":
                tot_big += 1
            lines.append(
                f"| {label} | {sheet} · `{cell}` | "
                f"{_fmt(ours, pct)} | {_fmt(theirs, pct)} | "
                f"{_fmt(delta_abs, pct)} | {delta_pct_str} | {flag} |"
            )
        lines.append("")

    lines.append(f"## Summary")
    lines.append("")
    lines.append(f"- ✓ matches (<1%): **{tot_match}**")
    lines.append(f"- ⚠ small (<5%):   **{tot_small}**")
    lines.append(f"- ❌ large (≥5%):   **{tot_big}**")
    lines.append("")
    return "\n".join(lines)


def main(tickers: list[str]):
    outdir = REPO_ROOT / "docs" / "experiments"
    outdir.mkdir(parents=True, exist_ok=True)
    for t in tickers:
        print(f"[{t}] comparing...")
        text = compare(t)
        out = outdir / f"ginzu_comparison_{t}.md"
        out.write_text(text)
        print(f"  wrote {out.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    args = sys.argv[1:] or list(COMPANIES.keys())
    main(args)
