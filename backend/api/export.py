"""Excel export endpoint — downloads visible sheet data as .xlsx."""

from __future__ import annotations

import io
import logging

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from .session_store import get_session

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api")

# Color fills matching frontend Tailwind classes
FILL_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # hypothesis
FILL_BLUE = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")    # financial
FILL_GREEN = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")   # calc
FILL_PURPLE = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")  # reference
FILL_GRAY = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")    # header
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")    # label

FILL_MAP = {
    "hypothesis": FILL_YELLOW,
    "financial": FILL_BLUE,
    "calc": FILL_GREEN,
    "reference": FILL_PURPLE,
    "header": FILL_GRAY,
    "label": FILL_WHITE,
    "hint": FILL_WHITE,
}


def _write_cell(ws, row: int, col: int, value, cell_type: str = "label"):
    """Write a cell with color fill."""
    cell = ws.cell(row=row, column=col, value=value)
    fill = FILL_MAP.get(cell_type, FILL_WHITE)
    cell.fill = fill
    return cell


def _build_input_sheet(session) -> Workbook:
    """Build the Input Sheet workbook from session data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Input Sheet"

    inp = session.inputs
    fin0 = inp.raw_financials[0] if inp.raw_financials else None
    fin1 = inp.raw_financials[1] if len(inp.raw_financials) > 1 else None
    macro = inp.macro_inputs
    ind = inp.industry_data
    va = inp.valuation_assumptions
    adj = inp.adjustment_inputs
    opt = inp.option_inputs

    r = 1

    # Section 1: Company Info
    _write_cell(ws, r, 1, "1. Company Information", "header")
    r += 1
    for label, val, typ in [
        ("Company Name", inp.company_name, "financial"),
        ("Ticker", inp.ticker, "financial"),
        ("Country", inp.country, "financial"),
        ("Reporting Currency", inp.reporting_currency, "financial"),
        ("Stock Price Currency", inp.stock_price_currency, "financial"),
        ("Industry", ind.industry_name, "reference"),
    ]:
        _write_cell(ws, r, 1, label, "label")
        _write_cell(ws, r, 2, val, typ)
        r += 1

    r += 1

    # Section 2: Financials
    _write_cell(ws, r, 1, "2. Base Year Financials", "header")
    r += 1
    headers = ["", "LTM"]
    for i, fin in enumerate(inp.raw_financials):
        headers.append(f"FY{fin.fiscal_year}")
    for c, h in enumerate(headers, 1):
        _write_cell(ws, r, c, h, "header")
    r += 1

    fin_rows = [
        ("Revenues", "revenues"), ("EBIT", "ebit"), ("EBITDA", "ebitda"),
        ("Net Income", "net_income"), ("Interest Expense", "interest_expense"),
        ("D&A", "d_a"), ("CapEx", "capex"), ("Non-cash WC", "noncash_wc"),
        ("Change in NWC", "change_in_noncash_wc"), ("Net Debt Issued", "net_debt_issued"),
        ("BV Equity", "bv_equity"), ("BV Debt", "bv_debt"),
    ]
    for label, attr in fin_rows:
        _write_cell(ws, r, 1, label, "label")
        col = 2
        # LTM column placeholder
        _write_cell(ws, r, col, getattr(fin0, attr, None) if fin0 else None, "calc")
        col += 1
        for fin in inp.raw_financials:
            _write_cell(ws, r, col, getattr(fin, attr, None), "financial")
            col += 1
        r += 1

    r += 1

    # Section 6: Tax
    _write_cell(ws, r, 1, "6. Tax Rates", "header")
    r += 1
    _write_cell(ws, r, 1, "Effective Tax Rate", "label")
    _write_cell(ws, r, 2, macro.tax_rate_effective, "financial")
    r += 1
    _write_cell(ws, r, 1, "Marginal Tax Rate", "label")
    _write_cell(ws, r, 2, macro.tax_rate_marginal, "hypothesis")
    r += 1

    r += 1

    # Section 8: Market
    _write_cell(ws, r, 1, "8. Market Numbers", "header")
    r += 1
    for label, val, typ in [
        ("Risk-free Rate", macro.risk_free_rate, "hypothesis"),
        ("ERP", macro.equity_risk_premium, "reference"),
        ("CRP", macro.country_risk_premium, "reference"),
    ]:
        _write_cell(ws, r, 1, label, "label")
        _write_cell(ws, r, 2, val, typ)
        r += 1

    return wb


# Map of sheet names to builder functions
_SHEET_BUILDERS = {
    "input-sheet": _build_input_sheet,
}


@router.get("/valuation/{session_id}/export/full-workbook")
def export_full_workbook(session_id: str):
    """Export the full valuation workbook (all 13 sheets) as Excel."""
    from fastapi.responses import Response
    from .export_workbook import generate_full_workbook
    session = get_session(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found")
    xlsx_bytes = generate_full_workbook(session.inputs, session.report)
    ticker = session.inputs.ticker
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{ticker}_valuation.xlsx"'},
    )


@router.get("/valuation/{session_id}/export/{sheet_name}")
def export_sheet(session_id: str, sheet_name: str):
    """Export a sheet as .xlsx file."""
    session = get_session(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found")

    builder = _SHEET_BUILDERS.get(sheet_name)
    if builder is None:
        # Generic fallback: export basic input data
        builder = _build_input_sheet

    wb = builder(session)

    # Write to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{session.inputs.ticker}_{sheet_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
