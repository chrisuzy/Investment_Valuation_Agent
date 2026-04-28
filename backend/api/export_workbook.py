"""
Export full valuation workbook — all 13 sheets matching frontend pages.

Each sheet mirrors the exact layout shown on the corresponding frontend page.
Cell types: blue=financial data, green=hypothesis/editable, gray=calculated, orange=reference.
Calculated cells use Excel formulas so the workbook works as a calculator.
"""

from __future__ import annotations

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

from engine.data_dictionary import CompanyValuationInput
from engine.orchestrator import ValuationReport


# ── Styles ────────────────────────────────────────────────────
BLUE_FILL = PatternFill("solid", fgColor="D6EAF8")       # financial data
GREEN_FILL = PatternFill("solid", fgColor="D5F5E3")      # hypothesis/editable
GRAY_FILL = PatternFill("solid", fgColor="F2F3F4")       # calculated
ORANGE_FILL = PatternFill("solid", fgColor="FEF5E7")     # reference
YELLOW_FILL = PatternFill("solid", fgColor="FEF9E7")     # hint
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NUM_FMT = '#,##0'
PCT_FMT = '0.00%'
DEC_FMT = '#,##0.00'


def _cell(ws, row, col, value, fill=None, font=None, fmt=None):
    """Write a styled cell."""
    c = ws.cell(row, col, value)
    c.border = THIN_BORDER
    if fill:
        c.fill = fill
    c.font = font or NORMAL_FONT
    if fmt:
        c.number_format = fmt
    return c


def _header_row(ws, row, headers, col_start=1):
    """Write a header row with dark background."""
    for i, h in enumerate(headers):
        _cell(ws, row, col_start + i, h, fill=HEADER_FILL, font=HEADER_FONT)


def _section_title(ws, row, title, num_cols=6):
    """Write a section title spanning multiple columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    c = _cell(ws, row, 1, title, font=Font(bold=True, size=11))
    c.fill = PatternFill("solid", fgColor="E8E8E8")


# ── Sheet Writers ─────────────────────────────────────────────

def _write_input_sheet(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport) -> dict:
    """Sheet 1: Input Sheet — mirrors frontend InputSheet.tsx.

    Returns a cell-position map (dict) so other sheets can build cross-sheet formulas.
    Keys are like 'revenues', 'ebit', 'tax_marginal', etc. Values are 'B13' style cell refs.
    """
    ws = wb.active
    ws.title = "Input Sheet"
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    for col_letter in 'CDEFGH':
        ws.column_dimensions[col_letter].width = 18

    fins = inp.raw_financials
    fin0 = fins[0] if fins else None
    adj = inp.adjustment_inputs
    macro = inp.macro_inputs
    ind = inp.industry_data
    va = inp.valuation_assumptions
    opt = inp.option_inputs

    # Cell position map — tracks where key values are for cross-sheet formulas
    pos = {}

    # === FIXED ROW LAYOUT ===
    # Row 1: Section header
    _section_title(ws, 1, "1. Company Information", 3)
    _cell(ws, 2, 1, "Date of Valuation", font=BOLD_FONT)
    _cell(ws, 2, 2, datetime.now().strftime("%Y-%m-%d"), fill=GREEN_FILL)
    _cell(ws, 3, 1, "Company Name", font=BOLD_FONT)
    _cell(ws, 3, 2, inp.company_name or "", fill=BLUE_FILL)
    _cell(ws, 4, 1, "Ticker", font=BOLD_FONT)
    _cell(ws, 4, 2, inp.ticker, fill=BLUE_FILL)
    _cell(ws, 5, 1, "Country", font=BOLD_FONT)
    _cell(ws, 5, 2, inp.country or "", fill=BLUE_FILL)
    _cell(ws, 6, 1, "Reported Currency", font=BOLD_FONT)
    _cell(ws, 6, 2, inp.reporting_currency or "", fill=BLUE_FILL)
    _cell(ws, 7, 1, "Stock Price Currency", font=BOLD_FONT)
    _cell(ws, 7, 2, inp.stock_price_currency or "", fill=BLUE_FILL)
    _cell(ws, 8, 1, "Industry (US)", font=BOLD_FONT)
    _cell(ws, 8, 2, ind.industry_name, fill=ORANGE_FILL)

    # Row 10: Section 2 header
    _section_title(ws, 10, "2. Base Year Financials", 8)
    # Row 11: column headers
    annual = fins[:5]
    headers = ["", "LTM"] + [str(f.fiscal_year) if f.fiscal_year > 100 else f"FY{f.fiscal_year}" for f in annual] + ["Source"]
    _header_row(ws, 11, headers)

    # Rows 12-20: financial line items (FIXED positions)
    fin_fields = [
        (12, "Revenues", "revenues"),
        (13, "EBIT (Operating Income)", "ebit"),
        (14, "EBITDA", "ebitda"),
        (15, "Net Income", "net_income"),
        (16, "Interest Expense", "interest_expense"),
        (17, "D&A", "d_a"),
        (18, "Capital Expenditures", "capex"),
        (19, "Book Value of Equity", "bv_equity"),
        (20, "Book Value of Debt", "bv_debt"),
    ]
    q_fins = inp.quarterly_financials or []
    for r, label, key in fin_fields:
        _cell(ws, r, 1, label, font=BOLD_FONT)
        # LTM column (col B) — formula if quarterly data, else copy from FY0
        if len(q_fins) >= 4:
            ltm_val = sum(getattr(q, key, 0) or 0 for q in q_fins[:4])
            _cell(ws, r, 2, ltm_val if ltm_val else None, fill=GRAY_FILL, fmt=NUM_FMT)
        else:
            # LTM = FY0 (col C) via formula
            _cell(ws, r, 2, f"=C{r}", fill=GRAY_FILL, fmt=NUM_FMT)
        # Annual columns (col C onwards) — raw input values
        for i, f in enumerate(annual):
            _cell(ws, r, 3 + i, getattr(f, key, None), fill=BLUE_FILL, fmt=NUM_FMT)
        _cell(ws, r, 3 + len(annual), "CIQ", fill=YELLOW_FILL)
        pos[key] = f"C{r}"  # FY0 column = C
        pos[f"{key}_ltm"] = f"B{r}"

    # Row 22: Section 3 Balance Sheet
    _section_title(ws, 22, "3. Balance Sheet", 4)
    bs_items = [
        (23, "Cash & Marketable Securities", "cash", fin0.cash_and_marketable_securities if fin0 else None, NUM_FMT),
        (24, "Cross Holdings", "cross_holdings", fin0.cross_holdings if fin0 else None, NUM_FMT),
        (25, "Minority Interests", "minority_interests", fin0.minority_interests if fin0 else None, NUM_FMT),
        (26, "Shares Outstanding", "shares", fin0.shares_outstanding if fin0 else None, NUM_FMT),
        (27, "Current Stock Price", "stock_price", fin0.stock_price if fin0 else None, DEC_FMT),
    ]
    for r, label, key, val, fmt in bs_items:
        _cell(ws, r, 1, label, font=BOLD_FONT)
        _cell(ws, r, 2, val, fill=BLUE_FILL, fmt=fmt)
        _cell(ws, r, 3, "CIQ", fill=YELLOW_FILL)
        pos[key] = f"B{r}"
    # Market Cap = Shares * Price (FORMULA)
    _cell(ws, 28, 1, "Market Cap", font=BOLD_FONT)
    _cell(ws, 28, 2, "=B26*B27", fill=GRAY_FILL, fmt=NUM_FMT)
    pos["market_cap"] = "B28"

    # Row 30: Section 4 Tax Rates
    _section_title(ws, 30, "4. Tax Rates", 4)
    _cell(ws, 31, 1, "Effective Tax Rate", font=BOLD_FONT)
    _cell(ws, 31, 2, macro.tax_rate_effective, fill=BLUE_FILL, fmt=PCT_FMT)
    pos["tax_effective"] = "B31"
    _cell(ws, 32, 1, "Marginal Tax Rate", font=BOLD_FONT)
    _cell(ws, 32, 2, macro.tax_rate_marginal, fill=GREEN_FILL, fmt=PCT_FMT)
    pos["tax_marginal"] = "B32"

    # Row 34: Section 5 Value Drivers
    _section_title(ws, 34, "5. Value Drivers", 4)
    _cell(ws, 35, 1, "Revenue Growth - Next Year", font=BOLD_FONT)
    _cell(ws, 35, 2, va.revenue_growth_next_year, fill=GREEN_FILL, fmt=PCT_FMT)
    pos["growth_yr1"] = "B35"
    _cell(ws, 36, 1, "Operating Margin - Next Year", font=BOLD_FONT)
    _cell(ws, 36, 2, f"=C13/C12", fill=GRAY_FILL, fmt=PCT_FMT)  # FORMULA: EBIT/Revenues
    pos["op_margin_current"] = "B36"
    _cell(ws, 37, 1, "Revenue Growth - Years 2-5", font=BOLD_FONT)
    _cell(ws, 37, 2, va.revenue_growth_years_2_5, fill=GREEN_FILL, fmt=PCT_FMT)
    pos["growth_yr2_5"] = "B37"
    _cell(ws, 38, 1, "Target Pre-tax Operating Margin", font=BOLD_FONT)
    _cell(ws, 38, 2, va.target_operating_margin, fill=GREEN_FILL, fmt=PCT_FMT)
    pos["target_margin"] = "B38"
    _cell(ws, 38, 3, f"Industry avg: {ind.pretax_operating_margin}", fill=YELLOW_FILL)
    _cell(ws, 39, 1, "Year of Convergence", font=BOLD_FONT)
    _cell(ws, 39, 2, va.margin_convergence_year, fill=GREEN_FILL)
    pos["convergence_year"] = "B39"
    _cell(ws, 40, 1, "Sales / Capital - Years 1-5", font=BOLD_FONT)
    _cell(ws, 40, 2, va.sales_to_capital_high, fill=GREEN_FILL, fmt=DEC_FMT)
    pos["sales_cap_high"] = "B40"
    _cell(ws, 41, 1, "Sales / Capital - Years 6-10", font=BOLD_FONT)
    _cell(ws, 41, 2, va.sales_to_capital_stable, fill=GREEN_FILL, fmt=DEC_FMT)
    pos["sales_cap_stable"] = "B41"

    # Row 43: Section 6 Market Numbers
    _section_title(ws, 43, "6. Market Numbers", 4)
    _cell(ws, 44, 1, "Risk-free Rate", font=BOLD_FONT)
    _cell(ws, 44, 2, macro.risk_free_rate, fill=GREEN_FILL, fmt=PCT_FMT)
    pos["risk_free"] = "B44"
    _cell(ws, 45, 1, "Equity Risk Premium", font=BOLD_FONT)
    _cell(ws, 45, 2, macro.equity_risk_premium, fill=ORANGE_FILL, fmt=PCT_FMT)
    pos["erp"] = "B45"
    _cell(ws, 46, 1, "Country Risk Premium", font=BOLD_FONT)
    _cell(ws, 46, 2, macro.country_risk_premium, fill=ORANGE_FILL, fmt=PCT_FMT)
    pos["crp"] = "B46"
    _cell(ws, 47, 1, "Initial WACC", font=BOLD_FONT)
    # FORMULA: cross-sheet ref to Cost of Capital WACC (will be set after CoC sheet is written)
    _cell(ws, 47, 2, "='Cost of Capital'!B28", fill=GRAY_FILL, fmt=PCT_FMT)
    pos["wacc"] = "B47"

    # Row 49: Section 7 Employee Options
    _section_title(ws, 49, "7. Employee Options", 4)
    _cell(ws, 50, 1, "Has Options Outstanding?", font=BOLD_FONT)
    _cell(ws, 50, 2, "Yes" if opt.has_options else "No", fill=GREEN_FILL)
    if opt.has_options:
        opts = [
            (51, "Number of Options", opt.number_of_options, NUM_FMT, "num_options"),
            (52, "Average Strike Price", opt.average_strike_price, DEC_FMT, "strike_price"),
            (53, "Average Maturity (years)", opt.average_maturity, DEC_FMT, "option_maturity"),
            (54, "Std Dev of Stock Price", opt.stock_price_std_dev, PCT_FMT, "stock_std_dev"),
            (55, "Dividend Yield", opt.dividend_yield, PCT_FMT, "div_yield"),
        ]
        for r, label, val, fmt, key in opts:
            _cell(ws, r, 1, label, font=BOLD_FONT)
            _cell(ws, r, 2, val, fill=GREEN_FILL, fmt=fmt)
            pos[key] = f"B{r}"

    return pos


def _write_answer_keys(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport):
    """Sheet 13: Answer Keys — mirrors AnswerKeys.tsx."""
    ws = wb.create_sheet("Answer Keys")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    fin0 = inp.raw_financials[0] if inp.raw_financials else None
    va = inp.valuation_assumptions
    ind = inp.industry_data
    macro = inp.macro_inputs
    coc = report.cost_of_capital
    dcf = report.dcf
    final_ = report.final

    row = 1

    sections = [
        ("Company Data", BLUE_FILL, [
            ("Company", inp.company_name or inp.ticker, None),
            ("Industry", ind.industry_name, None),
            ("Revenues", fin0.revenues if fin0 else None, NUM_FMT),
            ("EBIT", fin0.ebit if fin0 else None, NUM_FMT),
            ("Stock Price", fin0.stock_price if fin0 else None, DEC_FMT),
            ("Shares Outstanding", fin0.shares_outstanding if fin0 else None, NUM_FMT),
        ]),
        ("Key Assumptions", GREEN_FILL, [
            ("Revenue Growth (Yr 1)", va.revenue_growth_next_year, PCT_FMT),
            ("Revenue Growth (Yrs 2-5)", va.revenue_growth_years_2_5, PCT_FMT),
            ("Target Operating Margin", va.target_operating_margin, PCT_FMT),
            ("Convergence Year", va.margin_convergence_year, None),
            ("Risk-free Rate", macro.risk_free_rate, PCT_FMT),
            ("Failure Probability", va.failure_probability, PCT_FMT),
        ]),
        ("Computed Results", GRAY_FILL, [
            ("WACC", coc.wacc if coc else None, PCT_FMT),
            ("Levered Beta", coc.beta_l if coc else None, DEC_FMT),
            ("Terminal Value", dcf.terminal_value_firm if dcf else None, NUM_FMT),
            ("Value of Operating Assets", dcf.value_of_operating_assets if dcf else None, NUM_FMT),
            ("Value of Equity", dcf.value_of_equity if dcf else None, NUM_FMT),
            ("Value per Share", final_.value_per_share if final_ else None, DEC_FMT),
        ]),
        ("Industry Benchmarks", ORANGE_FILL, [
            ("Industry Beta (unlevered)", ind.beta_u, DEC_FMT),
            ("Industry D/E", ind.industry_d_e_ratio, DEC_FMT),
            ("Industry WACC", ind.wacc, PCT_FMT),
            ("Industry Operating Margin", ind.pretax_operating_margin, PCT_FMT),
            ("Industry Sales/Capital", ind.sales_to_capital, DEC_FMT),
        ]),
    ]

    for section_name, fill, items in sections:
        _section_title(ws, row, section_name, 2)
        row += 1
        _header_row(ws, row, ["Item", "Value"])
        row += 1
        for label, val, fmt in items:
            _cell(ws, row, 1, label)
            _cell(ws, row, 2, val, fill=fill, fmt=fmt)
            row += 1
        row += 1


def _write_trailing_twelve_month(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport):
    """Sheet 12: Trailing 12 Month — mirrors TrailingTwelveMonth.tsx."""
    ws = wb.create_sheet("Trailing 12 Month")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    fin0 = inp.raw_financials[0] if inp.raw_financials else None
    fin1 = inp.raw_financials[1] if len(inp.raw_financials) > 1 else None
    adj = inp.adjustment_inputs

    row = 1

    # Income Statement Items
    _section_title(ws, row, "Income Statement Items", 4)
    row += 1
    _header_row(ws, row, ["Item", "LTM / Most Recent", "Prior Annual", "TTM Computed"])
    row += 1

    is_fields = [
        ("Revenues", "revenues"), ("EBIT", "ebit"), ("EBITDA", "ebitda"),
        ("Net Income", "net_income"), ("Interest Expense", "interest_expense"),
        ("Capital Expenditures", "capex"), ("D&A", "d_a"),
    ]
    for label, key in is_fields:
        ltm = getattr(fin0, key, None) if fin0 else None
        annual = getattr(fin1, key, None) if fin1 else None
        _cell(ws, row, 1, label)
        _cell(ws, row, 2, ltm, fill=BLUE_FILL, fmt=NUM_FMT)
        _cell(ws, row, 3, annual, fill=BLUE_FILL, fmt=NUM_FMT)
        _cell(ws, row, 4, ltm, fill=GRAY_FILL, fmt=NUM_FMT)  # TTM = LTM for now
        row += 1
    row += 1

    # Balance Sheet Items
    _section_title(ws, row, "Balance Sheet Items (Most Recent)", 2)
    row += 1
    _header_row(ws, row, ["Item", "Value"])
    row += 1
    bs_fields = [
        ("Cash & Marketable Securities", fin0.cash_and_marketable_securities if fin0 else None),
        ("BV Equity", fin0.bv_equity if fin0 else None),
        ("BV Debt", fin0.bv_debt if fin0 else None),
        ("MV Equity", fin0.mv_equity if fin0 else None),
        ("Shares Outstanding", fin0.shares_outstanding if fin0 else None),
        ("Stock Price", fin0.stock_price if fin0 else None),
    ]
    for label, val in bs_fields:
        _cell(ws, row, 1, label)
        _cell(ws, row, 2, val, fill=BLUE_FILL, fmt=NUM_FMT)
        row += 1
    row += 1

    # Adjustment Inputs
    _section_title(ws, row, "Adjustment Inputs", 2)
    row += 1
    _cell(ws, row, 1, "Has R&D?")
    _cell(ws, row, 2, "Yes" if adj.has_r_and_d else "No", fill=GREEN_FILL)
    row += 1
    _cell(ws, row, 1, "R&D Expense (Current)")
    _cell(ws, row, 2, adj.r_and_d_expense_current, fill=GREEN_FILL, fmt=NUM_FMT)
    row += 1
    _cell(ws, row, 1, "Has Operating Leases?")
    _cell(ws, row, 2, "Yes" if adj.has_operating_leases else "No", fill=GREEN_FILL)
    row += 1
    _cell(ws, row, 1, "Lease Expense (Current)")
    _cell(ws, row, 2, adj.operating_lease_expense_current, fill=GREEN_FILL, fmt=NUM_FMT)


def _write_option_value(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport):
    """Sheet 11: Option Value — mirrors OptionValue.tsx (Black-Scholes)."""
    ws = wb.create_sheet("Option Value")
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

    import math

    opt = inp.option_inputs
    fin0 = inp.raw_financials[0] if inp.raw_financials else None
    macro = inp.macro_inputs
    final_ = report.final

    row = 1

    if not opt.has_options:
        _cell(ws, row, 1, "No employee options outstanding.", font=BOLD_FONT)
        return

    # Option Inputs
    _section_title(ws, row, "Option Inputs", 2)
    row += 1
    inputs_list = [
        ("Stock price", fin0.stock_price if fin0 else None, DEC_FMT),
        ("Average strike price", opt.average_strike_price, DEC_FMT),
        ("Average expiration (years)", opt.average_maturity, DEC_FMT),
        ("Standard deviation", opt.stock_price_std_dev, PCT_FMT),
        ("Dividend yield", opt.dividend_yield, PCT_FMT),
        ("Risk-free rate", macro.risk_free_rate, PCT_FMT),
        ("Number of options", opt.number_of_options, NUM_FMT),
        ("Number of shares", fin0.shares_outstanding if fin0 else None, NUM_FMT),
    ]
    for label, val, fmt in inputs_list:
        _cell(ws, row, 1, label, font=BOLD_FONT)
        _cell(ws, row, 2, val, fill=GREEN_FILL, fmt=fmt)
        row += 1
    row += 1

    # Black-Scholes
    _section_title(ws, row, "Dilution-Adjusted Black-Scholes", 2)
    row += 1

    S = None
    if fin0 and fin0.stock_price and fin0.shares_outstanding and opt.number_of_options:
        S = (fin0.stock_price * fin0.shares_outstanding + opt.average_strike_price * opt.number_of_options) / (fin0.shares_outstanding + opt.number_of_options)
    K = opt.average_strike_price
    t = opt.average_maturity
    sigma = opt.stock_price_std_dev
    r = macro.risk_free_rate
    y = opt.dividend_yield

    d1 = d2 = None
    if S and K > 0 and t > 0 and sigma > 0:
        d1 = (math.log(S / K) + (r - y + sigma**2 / 2) * t) / (sigma * math.sqrt(t))
        d2 = d1 - sigma * math.sqrt(t)

    def _norm_cdf(x):
        """Abramowitz & Stegun approximation."""
        a1, a2, a3, a4, a5, p = 0.254829592, -0.284496736, 1.421413741, -1.453152027, 1.061405429, 0.3275911
        sign = -1 if x < 0 else 1
        z = abs(x) / math.sqrt(2)
        tt = 1.0 / (1.0 + p * z)
        yy = 1.0 - ((((a5 * tt + a4) * tt + a3) * tt + a2) * tt + a1) * tt * math.exp(-z * z)
        return 0.5 * (1.0 + sign * yy)

    bs_rows = [
        ("Adjusted S (dilution-adjusted price)", S, DEC_FMT),
        ("K (strike price)", K, DEC_FMT),
        ("t (time to expiration)", t, DEC_FMT),
        ("sigma (std dev)", sigma, PCT_FMT),
        ("d1", d1, '0.0000'),
        ("N(d1)", _norm_cdf(d1) if d1 is not None else None, '0.0000'),
        ("d2", d2, '0.0000'),
        ("N(d2)", _norm_cdf(d2) if d2 is not None else None, '0.0000'),
        ("Value per option", final_.call_value_per_option if final_ else None, DEC_FMT),
        ("Total value of options", final_.value_of_all_options if final_ else None, NUM_FMT),
    ]
    for label, val, fmt in bs_rows:
        _cell(ws, row, 1, label, font=BOLD_FONT)
        _cell(ws, row, 2, val, fill=GRAY_FILL, fmt=fmt)
        row += 1


def _write_valuation_picture(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport):
    """Sheet 10: Valuation Picture — mirrors ValuationPicture.tsx."""
    ws = wb.create_sheet("Valuation Picture")
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 18

    fin0 = inp.raw_financials[0] if inp.raw_financials else None
    va = inp.valuation_assumptions
    dcf = report.dcf
    final_ = report.final
    coc = report.cost_of_capital
    macro = inp.macro_inputs

    base_rev = fin0.revenues if fin0 else 0
    term_rev = dcf.revenue_projections[-1] if dcf and dcf.revenue_projections else 0
    term_ebit = dcf.ebit_projections[-1] if dcf and dcf.ebit_projections else 0

    row = 1

    # Revenue Growth Path
    _section_title(ws, row, "Revenue Growth Path", 8)
    row += 1
    for col, label, val, fill, fmt in [
        (1, "Current revenues", base_rev, BLUE_FILL, NUM_FMT),
        (3, "Growth yr 1", va.revenue_growth_next_year, GREEN_FILL, PCT_FMT),
        (5, "Growth yrs 2-5", va.revenue_growth_years_2_5, GREEN_FILL, PCT_FMT),
        (7, "Terminal revenues", term_rev, GRAY_FILL, NUM_FMT),
    ]:
        _cell(ws, row, col, label, font=BOLD_FONT)
        _cell(ws, row, col + 1, val, fill=fill, fmt=fmt)
    row += 2

    # Margin Path
    _section_title(ws, row, "Operating Margin Path", 8)
    row += 1
    cur_margin = fin0.ebit / fin0.revenues if fin0 and fin0.revenues else 0
    for col, label, val, fill, fmt in [
        (1, "Current EBIT margin", cur_margin, GRAY_FILL, PCT_FMT),
        (3, "Target margin", va.target_operating_margin, GREEN_FILL, PCT_FMT),
        (5, "Converge year", va.margin_convergence_year, GREEN_FILL, None),
        (7, "Terminal EBIT", term_ebit, GRAY_FILL, NUM_FMT),
    ]:
        _cell(ws, row, col, label, font=BOLD_FONT)
        _cell(ws, row, col + 1, val, fill=fill, fmt=fmt)
    row += 2

    # Cost of Capital
    _section_title(ws, row, "Cost of Capital", 6)
    row += 1
    _cell(ws, row, 1, "Risk-free rate", font=BOLD_FONT)
    _cell(ws, row, 2, macro.risk_free_rate, fill=GREEN_FILL, fmt=PCT_FMT)
    _cell(ws, row, 3, "+ Beta x ERP", font=BOLD_FONT)
    _cell(ws, row, 4, macro.equity_risk_premium, fill=ORANGE_FILL, fmt=PCT_FMT)
    _cell(ws, row, 5, "WACC", font=BOLD_FONT)
    _cell(ws, row, 6, coc.wacc if coc else None, fill=GRAY_FILL, fmt=PCT_FMT)
    row += 2

    # Value Bridge
    _section_title(ws, row, "Value Bridge", 2)
    row += 1
    bridge = [
        ("PV(Cash Flows)", dcf.pv_cash_flows_sum if dcf else None, GRAY_FILL),
        ("+ PV(Terminal Value)", dcf.pv_terminal_value if dcf else None, GRAY_FILL),
        ("= Operating Assets", dcf.value_of_operating_assets if dcf else None, GRAY_FILL),
        ("- Debt", fin0.bv_debt if fin0 else None, BLUE_FILL),
        ("- Minority interests", fin0.minority_interests if fin0 else 0, BLUE_FILL),
        ("+ Cash", fin0.cash_and_marketable_securities if fin0 else None, BLUE_FILL),
        ("+ Non-operating assets", fin0.cross_holdings if fin0 else 0, BLUE_FILL),
        ("= Value of Equity", dcf.value_of_equity if dcf else None, GRAY_FILL),
        ("- Options", final_.value_of_all_options if final_ else 0, GRAY_FILL),
        ("/ Shares Outstanding", fin0.shares_outstanding if fin0 else None, BLUE_FILL),
        ("= Value per Share", final_.value_per_share if final_ else None, GRAY_FILL),
        ("Current Stock Price", fin0.stock_price if fin0 else None, BLUE_FILL),
    ]
    for label, val, fill in bridge:
        _cell(ws, row, 1, label, font=BOLD_FONT if label.startswith("=") else NORMAL_FONT)
        _cell(ws, row, 2, val, fill=fill, fmt=NUM_FMT if not isinstance(val, str) else None)
        row += 1


def _write_stories_to_numbers(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport):
    """Sheet 9: Stories to Numbers — mirrors StoriesToNumbers.tsx."""
    ws = wb.create_sheet("Stories to Numbers")
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15

    va = inp.valuation_assumptions
    ind = inp.industry_data
    fin0 = inp.raw_financials[0] if inp.raw_financials else None
    coc = report.cost_of_capital

    row = 1
    _header_row(ws, row, ["Narrative Question", "Value Driver", "Input Field", "Current Value", "Source"])
    row += 1

    stories = [
        ("How fast will the company grow revenues?", "Revenue growth (next year)", "revenue_growth_next_year", va.revenue_growth_next_year, "User Input", PCT_FMT, GREEN_FILL),
        ("How fast will it grow after year 1?", "Revenue growth (years 2-5)", "revenue_growth_years_2_5", va.revenue_growth_years_2_5, "User Input", PCT_FMT, GREEN_FILL),
        ("How profitable at maturity?", "Target pre-tax operating margin", "target_operating_margin", va.target_operating_margin, "User Input", PCT_FMT, GREEN_FILL),
        ("How long to reach target margin?", "Year of convergence", "margin_convergence_year", va.margin_convergence_year, "User Input", None, GREEN_FILL),
        ("How efficiently reinvest?", "Sales/Capital (years 1-5)", "sales_to_capital_high", va.sales_to_capital_high, "Calculated", DEC_FMT, GRAY_FILL),
        ("Reinvestment in stable growth?", "Sales/Capital (years 6-10)", "sales_to_capital_stable", va.sales_to_capital_stable, "Calculated", DEC_FMT, GRAY_FILL),
        ("How risky is this company?", "Cost of capital", "wacc", coc.wacc if coc else None, "Calculated", PCT_FMT, GRAY_FILL),
        ("What is the long-term growth rate?", "Stable growth rate", "stable_growth_rate", va.stable_growth_rate or inp.macro_inputs.risk_free_rate, "User Input", PCT_FMT, GREEN_FILL),
        ("Could this company fail?", "Probability of failure", "failure_probability", va.failure_probability, "User Input", PCT_FMT, GREEN_FILL),
        ("What are industry margins like?", "Industry pre-tax margin", "pretax_operating_margin", ind.pretax_operating_margin, "Industry/Market", PCT_FMT, ORANGE_FILL),
        ("Industry reinvestment?", "Industry sales/capital", "sales_to_capital", ind.sales_to_capital, "Industry/Market", DEC_FMT, ORANGE_FILL),
        ("What does the market say?", "Current stock price", "stock_price", fin0.stock_price if fin0 else None, "Industry/Market", DEC_FMT, ORANGE_FILL),
    ]

    for narrative, driver, field, value, source, fmt, fill in stories:
        _cell(ws, row, 1, narrative)
        _cell(ws, row, 2, driver)
        _cell(ws, row, 3, field, fill=YELLOW_FILL)
        _cell(ws, row, 4, value, fill=fill, fmt=fmt)
        _cell(ws, row, 5, source)
        row += 1


def _write_diagnostics(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport):
    """Sheet 8: Diagnostics — mirrors Diagnostics.tsx (6 diagnostic steps)."""
    ws = wb.create_sheet("Diagnostics")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18

    fin = inp.raw_financials
    va = inp.valuation_assumptions
    ind = inp.industry_data
    macro = inp.macro_inputs
    dcf = report.dcf
    fin0 = fin[0] if fin else None
    fin1 = fin[1] if len(fin) > 1 else None

    row = 1

    # Step 1: Revenue Growth
    _section_title(ws, row, "Step 1: Revenue Growth", 2)
    row += 1
    recent_growth = ((fin0.revenues - fin1.revenues) / abs(fin1.revenues)) if fin0 and fin1 and fin1.revenues else None
    _cell(ws, row, 1, "Industry avg revenue growth", font=BOLD_FONT)
    _cell(ws, row, 2, ind.revenue_growth, fill=ORANGE_FILL, fmt=PCT_FMT)
    row += 1
    _cell(ws, row, 1, "Most recent revenue growth", font=BOLD_FONT)
    _cell(ws, row, 2, recent_growth, fill=GRAY_FILL, fmt=PCT_FMT)
    row += 1
    _cell(ws, row, 1, "Forecast - Year 1 growth", font=BOLD_FONT)
    _cell(ws, row, 2, va.revenue_growth_next_year, fill=GREEN_FILL, fmt=PCT_FMT)
    row += 1
    _cell(ws, row, 1, "Forecast - Years 2-5 growth", font=BOLD_FONT)
    _cell(ws, row, 2, va.revenue_growth_years_2_5, fill=GREEN_FILL, fmt=PCT_FMT)
    row += 2

    # Step 2: Dollar Revenues
    _section_title(ws, row, "Step 2: Dollar Revenues", 2)
    row += 1
    _cell(ws, row, 1, "Base year revenues", font=BOLD_FONT)
    _cell(ws, row, 2, fin0.revenues if fin0 else None, fill=BLUE_FILL, fmt=NUM_FMT)
    row += 1
    if dcf and dcf.revenue_projections:
        for label, idx in [("Year 1 revenues", 0), ("Year 5 revenues", 4), ("Year 10 revenues", 9)]:
            val = dcf.revenue_projections[idx] if idx < len(dcf.revenue_projections) else None
            _cell(ws, row, 1, label, font=BOLD_FONT)
            _cell(ws, row, 2, val, fill=GRAY_FILL, fmt=NUM_FMT)
            row += 1
    row += 1

    # Step 3: Operating Margins
    _section_title(ws, row, "Step 3: Operating Margins", 2)
    row += 1
    current_margin = fin0.ebit / fin0.revenues if fin0 and fin0.revenues else None
    _cell(ws, row, 1, "Current operating margin", font=BOLD_FONT)
    _cell(ws, row, 2, current_margin, fill=GRAY_FILL, fmt=PCT_FMT)
    row += 1
    _cell(ws, row, 1, "Target operating margin", font=BOLD_FONT)
    _cell(ws, row, 2, va.target_operating_margin, fill=GREEN_FILL, fmt=PCT_FMT)
    row += 1
    _cell(ws, row, 1, "Industry pretax margin", font=BOLD_FONT)
    _cell(ws, row, 2, ind.pretax_operating_margin, fill=ORANGE_FILL, fmt=PCT_FMT)
    row += 1
    _cell(ws, row, 1, "Year of convergence", font=BOLD_FONT)
    _cell(ws, row, 2, va.margin_convergence_year, fill=GREEN_FILL)
    row += 2

    # Step 4: Reinvestment
    _section_title(ws, row, "Step 4: Reinvestment", 2)
    row += 1
    _cell(ws, row, 1, "Sales/capital years 1-5", font=BOLD_FONT)
    _cell(ws, row, 2, va.sales_to_capital_high, fill=GRAY_FILL, fmt=DEC_FMT)
    row += 1
    _cell(ws, row, 1, "Sales/capital years 6-10", font=BOLD_FONT)
    _cell(ws, row, 2, va.sales_to_capital_stable, fill=GRAY_FILL, fmt=DEC_FMT)
    row += 1
    _cell(ws, row, 1, "Industry sales/capital", font=BOLD_FONT)
    _cell(ws, row, 2, ind.sales_to_capital, fill=ORANGE_FILL, fmt=DEC_FMT)
    row += 2

    # Step 5: Risk
    _section_title(ws, row, "Step 5: Risk", 2)
    row += 1
    _cell(ws, row, 1, "Cost of capital (WACC)", font=BOLD_FONT)
    _cell(ws, row, 2, report.cost_of_capital.wacc if report.cost_of_capital else None, fill=GRAY_FILL, fmt=PCT_FMT)
    row += 1
    _cell(ws, row, 1, "Industry WACC", font=BOLD_FONT)
    _cell(ws, row, 2, ind.wacc, fill=ORANGE_FILL, fmt=PCT_FMT)
    row += 1
    _cell(ws, row, 1, "Risk-free rate", font=BOLD_FONT)
    _cell(ws, row, 2, macro.risk_free_rate, fill=GREEN_FILL, fmt=PCT_FMT)
    row += 2

    # Step 6: Value vs Price
    _section_title(ws, row, "Step 6: Value vs Price", 2)
    row += 1
    vps = report.final.value_per_share if report.final else None
    price = fin0.stock_price if fin0 else None
    _cell(ws, row, 1, "Value per share", font=BOLD_FONT)
    _cell(ws, row, 2, vps, fill=GRAY_FILL, fmt=DEC_FMT)
    row += 1
    _cell(ws, row, 1, "Stock price", font=BOLD_FONT)
    _cell(ws, row, 2, price, fill=BLUE_FILL, fmt=DEC_FMT)
    row += 1
    pct_val = price / vps if vps and price and vps != 0 else None
    _cell(ws, row, 1, "Price as % of value", font=BOLD_FONT)
    _cell(ws, row, 2, pct_val, fill=GRAY_FILL, fmt=PCT_FMT)


def _write_failure_rate(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport):
    """Sheet 7: Failure Rate — mirrors FailureRate.tsx."""
    ws = wb.create_sheet("Failure Rate")
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    va = inp.valuation_assumptions
    fail_prob = va.failure_probability
    proceeds_pct = va.distress_proceeds_pct
    op_assets = report.dcf.value_of_operating_assets if report.dcf else 0

    row = 1

    # Failure Probability Inputs
    _section_title(ws, row, "Failure Probability Inputs", 2)
    row += 1
    _cell(ws, row, 1, "Probability of failure", font=BOLD_FONT)
    _cell(ws, row, 2, fail_prob, fill=GREEN_FILL, fmt=PCT_FMT)
    row += 1
    _cell(ws, row, 1, "Proceeds as % of book value if failure", font=BOLD_FONT)
    _cell(ws, row, 2, proceeds_pct, fill=GREEN_FILL, fmt=PCT_FMT)
    row += 2

    # Impact on Valuation
    _section_title(ws, row, "Impact on Valuation", 2)
    row += 1
    _cell(ws, row, 1, "Value of operating assets (going concern)", font=BOLD_FONT)
    _cell(ws, row, 2, op_assets, fill=GRAY_FILL, fmt=NUM_FMT)
    row += 1
    _cell(ws, row, 1, "Probability of failure", font=BOLD_FONT)
    _cell(ws, row, 2, fail_prob, fill=GRAY_FILL, fmt=PCT_FMT)
    row += 1
    _cell(ws, row, 1, "Distress sale proceeds", font=BOLD_FONT)
    _cell(ws, row, 2, (op_assets or 0) * proceeds_pct, fill=GRAY_FILL, fmt=NUM_FMT)
    row += 1
    expected = (op_assets or 0) * (1 - fail_prob) + (op_assets or 0) * proceeds_pct * fail_prob
    _cell(ws, row, 1, "Expected operating asset value", font=Font(bold=True, size=11))
    _cell(ws, row, 2, expected, fill=GRAY_FILL, fmt=NUM_FMT, font=Font(bold=True, size=11))
    row += 2

    # Cumulative Default Rates
    _section_title(ws, row, "Cumulative Default Rates by Rating", 4)
    row += 1
    _header_row(ws, row, ["Bond Rating", "1-Year", "5-Year", "10-Year"])
    row += 1
    DEFAULT_RATES = [
        ("Aaa/AAA", 0.0001, 0.0012, 0.005), ("Aa2/AA", 0.0008, 0.005, 0.01),
        ("A1/A+", 0.002, 0.01, 0.02), ("A2/A", 0.005, 0.02, 0.04),
        ("Baa2/BBB", 0.012, 0.035, 0.07), ("Ba1/BB+", 0.02, 0.07, 0.11),
        ("Ba2/BB", 0.035, 0.11, 0.18), ("B1/B+", 0.05, 0.15, 0.25),
        ("B2/B", 0.065, 0.2, 0.32), ("B3/B-", 0.08, 0.25, 0.38),
        ("Caa/CCC", 0.12, 0.35, 0.5), ("Ca2/CC", 0.2, 0.45, 0.65),
        ("C2/C", 0.3, 0.6, 0.8), ("D2/D", 1.0, 1.0, 1.0),
    ]
    for rating, p1, p5, p10 in DEFAULT_RATES:
        _cell(ws, row, 1, rating, fill=ORANGE_FILL)
        _cell(ws, row, 2, p1, fill=ORANGE_FILL, fmt=PCT_FMT)
        _cell(ws, row, 3, p5, fill=ORANGE_FILL, fmt=PCT_FMT)
        _cell(ws, row, 4, p10, fill=ORANGE_FILL, fmt=PCT_FMT)
        row += 1


def _write_synthetic_rating(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport):
    """Sheet 6: Synthetic Rating — mirrors SyntheticRating.tsx."""
    ws = wb.create_sheet("Synthetic Rating")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    fin = inp.raw_financials[0] if inp.raw_financials else None
    ebit = report.adjusted.adjusted_ebit if report.adjusted else (fin.ebit if fin else 0)
    interest = fin.interest_expense if fin and fin.interest_expense else 1
    coverage = ebit / interest if interest > 0 else 999

    RATING_TABLE = [
        (-100, 0.5, "D2/D", 0.1486), (0.5, 0.8, "C2/C", 0.1286),
        (0.8, 1.25, "Ca2/CC", 0.1086), (1.25, 1.5, "Caa/CCC", 0.0886),
        (1.5, 2.0, "B3/B-", 0.0486), (2.0, 2.5, "B2/B", 0.0386),
        (2.5, 3.0, "B1/B+", 0.0336), (3.0, 3.5, "Ba2/BB", 0.0236),
        (3.5, 4.5, "Ba1/BB+", 0.0186), (4.5, 6.0, "Baa2/BBB", 0.0161),
        (6.0, 7.5, "A3/A-", 0.0136), (7.5, 9.5, "A2/A", 0.0111),
        (9.5, 12.5, "A1/A+", 0.0086), (12.5, 100, "Aaa/AAA", 0.0063),
    ]
    match = next((r for r in RATING_TABLE if coverage >= r[0] and coverage < r[1]), None)

    row = 1

    # Company Interest Coverage
    _section_title(ws, row, "Company Interest Coverage", 2)
    row += 1
    _cell(ws, row, 1, "EBIT (adjusted)", font=BOLD_FONT)
    _cell(ws, row, 2, ebit, fill=GRAY_FILL, fmt=NUM_FMT)
    row += 1
    _cell(ws, row, 1, "Interest expense", font=BOLD_FONT)
    _cell(ws, row, 2, interest, fill=BLUE_FILL, fmt=NUM_FMT)
    row += 1
    _cell(ws, row, 1, "Interest coverage ratio", font=Font(bold=True, size=11))
    _cell(ws, row, 2, coverage, fill=GRAY_FILL, fmt=DEC_FMT, font=Font(bold=True, size=11))
    row += 1
    _cell(ws, row, 1, "Estimated bond rating", font=Font(bold=True, size=11))
    _cell(ws, row, 2, match[2] if match else "N/A", fill=GRAY_FILL, font=Font(bold=True, size=11))
    row += 1
    _cell(ws, row, 1, "Estimated default spread", font=Font(bold=True, size=11))
    _cell(ws, row, 2, match[3] if match else 0, fill=GRAY_FILL, fmt=PCT_FMT, font=Font(bold=True, size=11))
    row += 2

    # Rating Lookup Table
    _section_title(ws, row, "Rating Lookup Table", 4)
    row += 1
    _header_row(ws, row, ["Min Coverage", "Max Coverage", "Rating", "Default Spread"])
    row += 1
    for mn, mx, rating, spread in RATING_TABLE:
        _cell(ws, row, 1, "< 0.5" if mn < 0 else mn, fill=ORANGE_FILL, fmt=DEC_FMT)
        _cell(ws, row, 2, "> 12.5" if mx >= 100 else mx, fill=ORANGE_FILL, fmt=DEC_FMT)
        _cell(ws, row, 3, rating, fill=ORANGE_FILL)
        _cell(ws, row, 4, spread, fill=ORANGE_FILL, fmt=PCT_FMT)
        row += 1


def _write_lease_converter(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport):
    """Sheet 5: Lease Converter — mirrors LeaseConverter.tsx."""
    ws = wb.create_sheet("Lease Converter")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20

    adj = inp.adjustment_inputs
    ind = inp.industry_data
    adjusted = report.adjusted
    commitments = adj.operating_lease_commitments
    cost_of_debt = ind.cost_of_debt_pretax or 0

    row = 1

    # Lease Commitments
    _section_title(ws, row, "Lease Commitments", 2)
    row += 1
    _cell(ws, row, 1, "Current year lease expense", font=BOLD_FONT)
    _cell(ws, row, 2, adj.operating_lease_expense_current, fill=GREEN_FILL, fmt=NUM_FMT)
    row += 1
    for i, c in enumerate(commitments):
        label = f"Year {i + 1}" if i < 5 else "Beyond Year 5"
        _cell(ws, row, 1, label, font=BOLD_FONT)
        _cell(ws, row, 2, c, fill=GREEN_FILL, fmt=NUM_FMT)
        row += 1
    row += 1

    # Discount Rate
    _section_title(ws, row, "Discount Rate", 2)
    row += 1
    _cell(ws, row, 1, "Pre-tax cost of debt", font=BOLD_FONT)
    _cell(ws, row, 2, cost_of_debt, fill=ORANGE_FILL, fmt=PCT_FMT)
    row += 2

    # PV Calculation
    _section_title(ws, row, "PV of Lease Commitments", 4)
    row += 1
    _header_row(ws, row, ["Year", "Commitment", "Discount Factor", "PV of Commitment"])
    row += 1

    total_pv = 0.0
    if cost_of_debt > 0:
        for i in range(min(len(commitments), 5)):
            c = commitments[i]
            df = 1 / ((1 + cost_of_debt) ** (i + 1))
            pv = c * df
            total_pv += pv
            _cell(ws, row, 1, f"Year {i + 1}")
            _cell(ws, row, 2, c, fill=GRAY_FILL, fmt=NUM_FMT)
            _cell(ws, row, 3, df, fill=GRAY_FILL, fmt='0.0000')
            _cell(ws, row, 4, pv, fill=GRAY_FILL, fmt=NUM_FMT)
            row += 1

        # Beyond year 5
        if len(commitments) > 5 and commitments[5] > 0:
            beyond = commitments[5]
            yr5_commit = commitments[4] if len(commitments) >= 5 else 0
            n_add = max(1, round(beyond / yr5_commit)) if yr5_commit > 0 else 1
            annual_amt = beyond / n_add
            for j in range(n_add):
                year = 6 + j
                df = 1 / ((1 + cost_of_debt) ** year)
                pv = annual_amt * df
                total_pv += pv
                label = f"Year {year}" if n_add == 1 else (f"Year 6 (of {n_add} beyond)" if j == 0 else f"Year {year}")
                _cell(ws, row, 1, label)
                _cell(ws, row, 2, annual_amt, fill=GRAY_FILL, fmt=NUM_FMT)
                _cell(ws, row, 3, df, fill=GRAY_FILL, fmt='0.0000')
                _cell(ws, row, 4, pv, fill=GRAY_FILL, fmt=NUM_FMT)
                row += 1

    _cell(ws, row, 1, "Total: PV of Operating Leases", font=BOLD_FONT)
    _cell(ws, row, 4, adjusted.pv_of_operating_leases if adjusted else total_pv, fill=GRAY_FILL, fmt=NUM_FMT, font=BOLD_FONT)
    row += 2

    # Adjustments Summary
    _section_title(ws, row, "Adjustments Summary", 2)
    row += 1
    for label, val in [
        ("Debt value of leases", adjusted.pv_of_operating_leases if adjusted else None),
        ("Adjusted BV equity", adjusted.adjusted_bv_equity if adjusted else None),
        ("Adjusted MV debt", adjusted.adjusted_mv_debt if adjusted else None),
    ]:
        _cell(ws, row, 1, label, font=BOLD_FONT)
        _cell(ws, row, 2, val, fill=GRAY_FILL, fmt=NUM_FMT)
        row += 1


def _write_rd_converter(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport):
    """Sheet 4: R&D Converter — mirrors RDConverter.tsx."""
    ws = wb.create_sheet("R&D Converter")
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22

    adj = inp.adjustment_inputs
    adjusted = report.adjusted
    n = adj.amortization_period_n
    current_rd = adj.r_and_d_expense_current
    past_rd = adj.r_and_d_expense_past

    row = 1

    # Section 1: R&D Expenses
    _section_title(ws, row, "R&D Expenses", 2)
    row += 1
    _cell(ws, row, 1, "Amortization Period (years)", font=BOLD_FONT)
    _cell(ws, row, 2, n, fill=GREEN_FILL)
    row += 1
    _cell(ws, row, 1, "Current Year R&D", font=BOLD_FONT)
    _cell(ws, row, 2, current_rd, fill=GREEN_FILL, fmt=NUM_FMT)
    row += 1
    for i, expense in enumerate(past_rd):
        _cell(ws, row, 1, f"t-{i + 1}", font=BOLD_FONT)
        _cell(ws, row, 2, expense, fill=GREEN_FILL, fmt=NUM_FMT)
        row += 1
    row += 1

    # Section 2: Amortization Schedule
    _section_title(ws, row, "Amortization Schedule", 5)
    row += 1
    _header_row(ws, row, ["Year", "R&D Expense", "Unamortized %", "Unamortized Amount", "Amortization This Year"])
    row += 1

    # Current year row
    _cell(ws, row, 1, "Current")
    _cell(ws, row, 2, current_rd, fill=GREEN_FILL, fmt=NUM_FMT)
    _cell(ws, row, 3, 1.0, fill=GRAY_FILL, fmt='0.00')
    _cell(ws, row, 4, current_rd, fill=GRAY_FILL, fmt=NUM_FMT)
    _cell(ws, row, 5, None, fill=GRAY_FILL)
    row += 1

    total_unamortized = current_rd
    total_amortization = 0.0
    for i, expense in enumerate(past_rd):
        years_since = i + 1
        unamortized_frac = max(0, (n - years_since) / n)
        unamortized_amt = expense * unamortized_frac
        amort_this_yr = expense / n if years_since <= n else 0
        total_unamortized += unamortized_amt
        total_amortization += amort_this_yr

        _cell(ws, row, 1, f"t-{years_since}")
        _cell(ws, row, 2, expense, fill=GREEN_FILL, fmt=NUM_FMT)
        _cell(ws, row, 3, unamortized_frac, fill=GRAY_FILL, fmt='0.00')
        _cell(ws, row, 4, unamortized_amt, fill=GRAY_FILL, fmt=NUM_FMT)
        _cell(ws, row, 5, amort_this_yr, fill=GRAY_FILL, fmt=NUM_FMT)
        row += 1

    # Totals
    _cell(ws, row, 1, "Total", font=BOLD_FONT)
    _cell(ws, row, 4, total_unamortized, fill=GRAY_FILL, fmt=NUM_FMT, font=BOLD_FONT)
    _cell(ws, row, 5, total_amortization, fill=GRAY_FILL, fmt=NUM_FMT, font=BOLD_FONT)
    row += 2

    # Section 3: Summary
    _section_title(ws, row, "Summary", 2)
    row += 1
    for label, val in [
        ("Value of Research Asset", adjusted.value_of_research_asset if adjusted else None),
        ("Unamortized R&D", adjusted.unamortized_r_and_d if adjusted else None),
        ("Amortization of R&D", adjusted.amortization_r_and_d if adjusted else None),
        ("Adjusted EBIT", adjusted.adjusted_ebit if adjusted else None),
    ]:
        _cell(ws, row, 1, label, font=BOLD_FONT)
        _cell(ws, row, 2, val, fill=GRAY_FILL, fmt=NUM_FMT)
        row += 1


def _write_cost_of_capital(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport, pos: dict):
    """Sheet 3: Cost of Capital — ALL computed cells are Excel formulas.

    Fixed layout: WACC is at B28 (referenced by Input Sheet B47).
    """
    ws = wb.create_sheet("Cost of Capital")
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

    ind = inp.industry_data
    IS = "'Input Sheet'"  # shorthand for cross-sheet refs

    # Row 1: Beta Estimation
    _section_title(ws, 1, "Beta Estimation", 2)
    _cell(ws, 2, 1, "Beta estimation source", font=BOLD_FONT)
    _cell(ws, 2, 2, "Bottom up", fill=GREEN_FILL)
    _cell(ws, 3, 1, "Unlevered beta", font=BOLD_FONT)
    _cell(ws, 3, 2, ind.beta_u, fill=ORANGE_FILL, fmt='0.0000')  # input from Damodaran
    _cell(ws, 4, 1, "Unlevered beta (corrected for cash)", font=BOLD_FONT)
    _cell(ws, 4, 2, ind.beta_u_corrected_for_cash, fill=ORANGE_FILL, fmt='0.0000')

    # Row 6: Debt Inputs
    _section_title(ws, 6, "Debt Inputs", 2)
    _cell(ws, 7, 1, "BV debt", font=BOLD_FONT)
    _cell(ws, 7, 2, f"={IS}!{pos.get('bv_debt','B20')}", fill=BLUE_FILL, fmt=NUM_FMT)  # cross-sheet
    _cell(ws, 8, 1, "Interest expense", font=BOLD_FONT)
    _cell(ws, 8, 2, f"={IS}!{pos.get('interest_expense','C16')}", fill=BLUE_FILL, fmt=NUM_FMT)
    _cell(ws, 9, 1, "Pre-tax cost of debt", font=BOLD_FONT)
    _cell(ws, 9, 2, ind.cost_of_debt_pretax, fill=ORANGE_FILL, fmt=PCT_FMT)  # Damodaran input

    # Row 11: Capital Structure — ALL FORMULAS
    _section_title(ws, 11, "Capital Structure", 2)
    _cell(ws, 12, 1, "Market Cap (equity)", font=BOLD_FONT)
    _cell(ws, 12, 2, f"={IS}!{pos.get('market_cap','B28')}", fill=BLUE_FILL, fmt=NUM_FMT)
    _cell(ws, 13, 1, "D/E ratio", font=BOLD_FONT)
    _cell(ws, 13, 2, "=B7/B12", fill=GRAY_FILL, fmt='0.0000')  # BV Debt / Market Cap
    _cell(ws, 14, 1, "Levered beta", font=BOLD_FONT)
    # beta_L = beta_U * (1 + (1 - tax) * D/E)
    _cell(ws, 14, 2, f"=B3*(1+(1-{IS}!{pos.get('tax_marginal','B32')})*B13)", fill=GRAY_FILL, fmt='0.0000')
    _cell(ws, 15, 1, "Weight of equity", font=BOLD_FONT)
    _cell(ws, 15, 2, "=B12/(B12+B7)", fill=GRAY_FILL, fmt=PCT_FMT)  # MktCap / (MktCap+Debt)
    _cell(ws, 16, 1, "Weight of debt", font=BOLD_FONT)
    _cell(ws, 16, 2, "=B7/(B12+B7)", fill=GRAY_FILL, fmt=PCT_FMT)

    # Row 18: Cost of Capital — ALL FORMULAS
    _section_title(ws, 18, "Cost of Capital", 2)
    _cell(ws, 19, 1, "Risk-free rate", font=BOLD_FONT)
    _cell(ws, 19, 2, f"={IS}!{pos.get('risk_free','B44')}", fill=GREEN_FILL, fmt=PCT_FMT)
    _cell(ws, 20, 1, "Equity risk premium (ERP)", font=BOLD_FONT)
    _cell(ws, 20, 2, f"={IS}!{pos.get('erp','B45')}", fill=ORANGE_FILL, fmt=PCT_FMT)
    _cell(ws, 21, 1, "Country risk premium", font=BOLD_FONT)
    _cell(ws, 21, 2, f"={IS}!{pos.get('crp','B46')}", fill=ORANGE_FILL, fmt=PCT_FMT)
    _cell(ws, 22, 1, "Cost of equity", font=BOLD_FONT)
    # CoE = Rf + Beta_L * ERP
    _cell(ws, 22, 2, "=B19+B14*B20", fill=GRAY_FILL, fmt=PCT_FMT)
    _cell(ws, 23, 1, "Cost of debt (pre-tax)", font=BOLD_FONT)
    _cell(ws, 23, 2, "=B9", fill=GRAY_FILL, fmt=PCT_FMT)  # ref to Damodaran input
    _cell(ws, 24, 1, "Cost of debt (after-tax)", font=BOLD_FONT)
    # CoD_at = CoD * (1 - t)
    _cell(ws, 24, 2, f"=B23*(1-{IS}!{pos.get('tax_marginal','B32')})", fill=GRAY_FILL, fmt=PCT_FMT)
    _cell(ws, 25, 1, "Marginal tax rate", font=BOLD_FONT)
    _cell(ws, 25, 2, f"={IS}!{pos.get('tax_marginal','B32')}", fill=GREEN_FILL, fmt=PCT_FMT)
    # WACC = CoE * We + CoD_at * Wd
    _cell(ws, 28, 1, "WACC", font=Font(bold=True, size=11))
    _cell(ws, 28, 2, "=B22*B15+B24*B16", fill=GRAY_FILL, fmt=PCT_FMT, font=Font(bold=True, size=11))
    pos["coc_wacc"] = "B28"

    # Row 30: Industry Comparison — static reference data
    _section_title(ws, 30, "Industry Comparison", 2)
    ind_items = [
        (31, "Industry", ind.industry_name, None),
        (32, "Industry D/E ratio", ind.industry_d_e_ratio, DEC_FMT),
        (33, "Industry effective tax rate", ind.industry_effective_tax_rate, PCT_FMT),
        (34, "Industry WACC", ind.wacc, PCT_FMT),
        (35, "Industry cost of equity", ind.cost_of_equity, PCT_FMT),
    ]
    for r, label, val, fmt in ind_items:
        _cell(ws, r, 1, label, font=BOLD_FONT)
        _cell(ws, r, 2, val, fill=ORANGE_FILL, fmt=fmt)


def _write_valuation_output(wb: openpyxl.Workbook, inp: CompanyValuationInput, report: ValuationReport, pos: dict):
    """Sheet 2: Valuation Output — ALL computed cells are Excel formulas.

    Fixed layout:
    Row 1: section title
    Row 2: headers (B=Base year, C-L=Year 1-10, M=Terminal year, N=Terminal Value)
    Row 3: Revenue growth rate    Row 4: Revenues
    Row 5: EBIT margin            Row 6: EBIT
    Row 7: Tax rate                Row 8: EBIT(1-t)
    Row 9: - Reinvestment          Row 10: FCFF
    Row 11: Cost of capital        Row 12: Cumulated discount factor
    Row 13: PV(FCFF)
    Row 15+: Value Bridge
    """
    ws = wb.create_sheet("Valuation Output")
    ws.column_dimensions['A'].width = 35
    for col_letter in 'BCDEFGHIJKLMN':
        ws.column_dimensions[col_letter].width = 15

    from openpyxl.utils import get_column_letter as gcl
    IS = "'Input Sheet'"
    COC = "'Cost of Capital'"

    # Row 1: Section title
    _section_title(ws, 1, "Projection Table", 14)

    # Row 2: Headers — B=Base year, C=Year 1 ... L=Year 10, M=Terminal year
    headers = ["", "Base year"] + [f"Year {i}" for i in range(1, 11)] + ["Terminal year"]
    _header_row(ws, 2, headers)

    # --- Row 3: Revenue growth rate ---
    _cell(ws, 3, 1, "Revenue growth rate", font=BOLD_FONT)
    _cell(ws, 3, 2, None, fill=GRAY_FILL)  # base year: no growth
    # Year 1: from Input Sheet growth_yr1
    _cell(ws, 3, 3, f"={IS}!{pos.get('growth_yr1','B35')}", fill=GREEN_FILL, fmt=PCT_FMT)
    # Years 2-5: from Input Sheet growth_yr2_5
    for col in range(4, 8):  # D-G = years 2-5
        _cell(ws, 3, col, f"={IS}!{pos.get('growth_yr2_5','B37')}", fill=GREEN_FILL, fmt=PCT_FMT)
    # Years 6-10: linearly converge to stable growth (risk-free rate)
    # Simplified: use risk-free rate as stable growth
    for col in range(8, 13):  # H-L = years 6-10
        yr = col - 2  # year number (6-10)
        # Linear interpolation: g_yr5 + (yr-5)/5 * (g_stable - g_yr5)
        _cell(ws, 3, col, f"=G3+(({yr}-5)/5)*({IS}!{pos.get('risk_free','B44')}-G3)", fill=GRAY_FILL, fmt=PCT_FMT)
    # Terminal: stable growth = risk-free rate
    _cell(ws, 3, 13, f"={IS}!{pos.get('risk_free','B44')}", fill=GRAY_FILL, fmt=PCT_FMT)

    # --- Row 4: Revenues ---
    _cell(ws, 4, 1, "Revenues", font=BOLD_FONT)
    _cell(ws, 4, 2, f"={IS}!{pos.get('revenues','C12')}", fill=BLUE_FILL, fmt=NUM_FMT)  # base year from Input
    # Year 1..10: Rev_t = Rev_{t-1} * (1 + growth_t)
    for col in range(3, 13):  # C-L
        prev = gcl(col - 1)
        _cell(ws, 4, col, f"={prev}4*(1+{gcl(col)}3)", fill=GRAY_FILL, fmt=NUM_FMT)
    # Terminal: Rev_terminal = Rev_yr10 * (1 + g_stable)
    _cell(ws, 4, 13, "=L4*(1+M3)", fill=GRAY_FILL, fmt=NUM_FMT)

    # --- Row 5: EBIT margin ---
    _cell(ws, 5, 1, "EBIT margin", font=BOLD_FONT)
    # Base year: EBIT/Revenue from Input Sheet
    _cell(ws, 5, 2, f"={IS}!{pos.get('ebit','C13')}/{IS}!{pos.get('revenues','C12')}", fill=GRAY_FILL, fmt=PCT_FMT)
    # Year 1-5: linear convergence from current margin to target margin
    for col in range(3, 8):  # C-G = years 1-5
        yr = col - 2
        conv = pos.get('convergence_year', 'B39')
        _cell(ws, 5, col, f"=B5+({yr}/{IS}!{conv})*({IS}!{pos.get('target_margin','B38')}-B5)", fill=GRAY_FILL, fmt=PCT_FMT)
    # Years 6-10 and terminal: at target margin
    for col in range(8, 14):  # H-M
        _cell(ws, 5, col, f"={IS}!{pos.get('target_margin','B38')}", fill=GRAY_FILL, fmt=PCT_FMT)

    # --- Row 6: EBIT = Revenues × Margin ---
    _cell(ws, 6, 1, "EBIT", font=BOLD_FONT)
    _cell(ws, 6, 2, f"={IS}!{pos.get('ebit','C13')}", fill=BLUE_FILL, fmt=NUM_FMT)
    for col in range(3, 14):  # C-M
        _cell(ws, 6, col, f"={gcl(col)}4*{gcl(col)}5", fill=GRAY_FILL, fmt=NUM_FMT)

    # --- Row 7: Tax rate ---
    _cell(ws, 7, 1, "Tax rate", font=BOLD_FONT)
    for col in range(2, 14):  # B-M: all reference Input Sheet marginal tax
        _cell(ws, 7, col, f"={IS}!{pos.get('tax_marginal','B32')}", fill=GREEN_FILL, fmt=PCT_FMT)

    # --- Row 8: EBIT(1-t) = EBIT × (1 - tax) ---
    _cell(ws, 8, 1, "EBIT(1-t)", font=BOLD_FONT)
    for col in range(2, 14):
        _cell(ws, 8, col, f"={gcl(col)}6*(1-{gcl(col)}7)", fill=GRAY_FILL, fmt=NUM_FMT)

    # --- Row 9: Reinvestment = ΔRevenue / Sales-to-Capital ---
    _cell(ws, 9, 1, "- Reinvestment", font=BOLD_FONT)
    _cell(ws, 9, 2, None, fill=GRAY_FILL)  # base year: no reinvestment
    # Years 1-5: use sales_cap_high
    for col in range(3, 8):
        prev = gcl(col - 1)
        _cell(ws, 9, col, f"=({gcl(col)}4-{prev}4)/{IS}!{pos.get('sales_cap_high','B40')}", fill=GRAY_FILL, fmt=NUM_FMT)
    # Years 6-10: use sales_cap_stable
    for col in range(8, 13):
        prev = gcl(col - 1)
        _cell(ws, 9, col, f"=({gcl(col)}4-{prev}4)/{IS}!{pos.get('sales_cap_stable','B41')}", fill=GRAY_FILL, fmt=NUM_FMT)
    # Terminal reinvestment = g_stable / ROIC_stable * EBIT(1-t)_terminal
    # Simplified: terminal reinvestment rate = g / WACC (if ROIC = WACC in stable)
    _cell(ws, 9, 13, f"=M3/{COC}!B28*M8", fill=GRAY_FILL, fmt=NUM_FMT)

    # --- Row 10: FCFF = EBIT(1-t) - Reinvestment ---
    _cell(ws, 10, 1, "FCFF", font=BOLD_FONT)
    _cell(ws, 10, 2, None, fill=GRAY_FILL)
    for col in range(3, 14):  # C-M
        _cell(ws, 10, col, f"={gcl(col)}8-{gcl(col)}9", fill=GRAY_FILL, fmt=NUM_FMT)

    # --- Row 11: Cost of capital = WACC from Cost of Capital sheet ---
    _cell(ws, 11, 1, "Cost of capital", font=BOLD_FONT)
    _cell(ws, 11, 2, None, fill=GRAY_FILL)
    for col in range(3, 14):
        _cell(ws, 11, col, f"={COC}!B28", fill=GRAY_FILL, fmt=PCT_FMT)

    # --- Row 12: Cumulated discount factor = 1/(1+WACC)^t ---
    _cell(ws, 12, 1, "Cumulated discount factor", font=BOLD_FONT)
    _cell(ws, 12, 2, None, fill=GRAY_FILL)
    for col in range(3, 13):  # C-L = years 1-10
        yr = col - 2
        _cell(ws, 12, col, f"=1/(1+{gcl(col)}11)^{yr}", fill=GRAY_FILL, fmt='0.0000')

    # --- Row 13: PV(FCFF) = FCFF × discount factor ---
    _cell(ws, 13, 1, "PV(FCFF)", font=BOLD_FONT)
    _cell(ws, 13, 2, None, fill=GRAY_FILL)
    for col in range(3, 13):
        _cell(ws, 13, col, f"={gcl(col)}10*{gcl(col)}12", fill=GRAY_FILL, fmt=NUM_FMT)
    # Terminal value = Terminal FCFF / (WACC - g_stable), in col M
    _cell(ws, 13, 13, "=M10/(M11-M3)", fill=GRAY_FILL, fmt=NUM_FMT)
    # Col N: terminal value label
    _cell(ws, 2, 14, "Terminal Value", fill=HEADER_FILL, font=HEADER_FONT)
    _cell(ws, 13, 14, "=M13", fill=GRAY_FILL, fmt=NUM_FMT)

    # === Row 15+: Value Bridge (FIXED ROWS) ===
    _section_title(ws, 15, "Value Bridge", 3)
    # Row 16: PV of cash flows = SUM of PV(FCFF) years 1-10
    _cell(ws, 16, 1, "PV(cash flows, years 1-10)", font=BOLD_FONT)
    _cell(ws, 16, 2, "=SUM(C13:L13)", fill=GRAY_FILL, fmt=NUM_FMT)
    # Row 17: PV of terminal value = terminal_value * discount_factor_yr10
    _cell(ws, 17, 1, "PV(terminal value)", font=BOLD_FONT)
    _cell(ws, 17, 2, "=M13*L12", fill=GRAY_FILL, fmt=NUM_FMT)
    # Row 18: Value of operating assets = PV_cf + PV_tv
    _cell(ws, 18, 1, "Value of operating assets", font=Font(bold=True, size=11))
    _cell(ws, 18, 2, "=B16+B17", fill=GRAY_FILL, fmt=NUM_FMT, font=Font(bold=True, size=11))
    # Row 19: - Debt (from Input Sheet)
    _cell(ws, 19, 1, "- Debt", font=BOLD_FONT)
    _cell(ws, 19, 2, f"={IS}!{pos.get('bv_debt','C20')}", fill=BLUE_FILL, fmt=NUM_FMT)
    # Row 20: - Minority interests
    _cell(ws, 20, 1, "- Minority interests", font=BOLD_FONT)
    _cell(ws, 20, 2, f"={IS}!{pos.get('minority_interests','B25')}", fill=BLUE_FILL, fmt=NUM_FMT)
    # Row 21: + Cash
    _cell(ws, 21, 1, "+ Cash", font=BOLD_FONT)
    _cell(ws, 21, 2, f"={IS}!{pos.get('cash','B23')}", fill=BLUE_FILL, fmt=NUM_FMT)
    # Row 22: + Cross holdings
    _cell(ws, 22, 1, "+ Non-operating assets (cross holdings)", font=BOLD_FONT)
    _cell(ws, 22, 2, f"={IS}!{pos.get('cross_holdings','B24')}", fill=BLUE_FILL, fmt=NUM_FMT)
    # Row 23: Value of equity = OpAssets - Debt - Minority + Cash + CrossHoldings
    _cell(ws, 23, 1, "Value of equity", font=Font(bold=True, size=11))
    _cell(ws, 23, 2, "=B18-B19-B20+B21+B22", fill=GRAY_FILL, fmt=NUM_FMT, font=Font(bold=True, size=11))
    # Row 24: - Value of options (cross-sheet to Option Value, or 0)
    _cell(ws, 24, 1, "- Value of options", font=BOLD_FONT)
    _cell(ws, 24, 2, "='Option Value'!B20", fill=GRAY_FILL, fmt=NUM_FMT)
    # Row 25: Equity in common stock = Equity - Options
    _cell(ws, 25, 1, "Value of equity in common stock", font=Font(bold=True, size=11))
    _cell(ws, 25, 2, "=B23-B24", fill=GRAY_FILL, fmt=NUM_FMT, font=Font(bold=True, size=11))
    # Row 26: Number of shares
    _cell(ws, 26, 1, "Number of shares", font=BOLD_FONT)
    _cell(ws, 26, 2, f"={IS}!{pos.get('shares','B26')}", fill=BLUE_FILL, fmt=NUM_FMT)
    # Row 27: Value per share = Equity / Shares
    _cell(ws, 27, 1, "Estimated value per share", font=Font(bold=True, size=11))
    _cell(ws, 27, 2, "=B25/B26", fill=GRAY_FILL, fmt=DEC_FMT, font=Font(bold=True, size=11))
    # Row 28: Current price
    _cell(ws, 28, 1, "Current price", font=BOLD_FONT)
    _cell(ws, 28, 2, f"={IS}!{pos.get('stock_price','B27')}", fill=BLUE_FILL, fmt=DEC_FMT)
    # Row 29: Price as % of value
    _cell(ws, 29, 1, "Price as % of value", font=BOLD_FONT)
    _cell(ws, 29, 2, "=B28/B27", fill=GRAY_FILL, fmt=PCT_FMT)


def generate_full_workbook(inp: CompanyValuationInput, report: ValuationReport) -> bytes:
    """Generate the full valuation workbook and return as bytes."""
    wb = openpyxl.Workbook()

    # Sheet 1: Input Sheet — returns cell position map for cross-sheet formulas
    pos = _write_input_sheet(wb, inp, report)

    # Sheet 2: Valuation Output (DCF) — needs pos for cross-sheet refs
    _write_valuation_output(wb, inp, report, pos)

    # Sheet 3: Cost of Capital (needs pos for cross-sheet refs)
    _write_cost_of_capital(wb, inp, report, pos)

    # Sheet 4: R&D Converter
    _write_rd_converter(wb, inp, report)

    # Sheet 5: Lease Converter
    _write_lease_converter(wb, inp, report)

    # Sheet 6: Synthetic Rating
    _write_synthetic_rating(wb, inp, report)

    # Sheet 7: Failure Rate
    _write_failure_rate(wb, inp, report)

    # Sheet 8: Diagnostics
    _write_diagnostics(wb, inp, report)

    # Sheet 9: Stories to Numbers
    _write_stories_to_numbers(wb, inp, report)

    # Sheet 10: Valuation Picture
    _write_valuation_picture(wb, inp, report)

    # Sheet 11: Option Value
    _write_option_value(wb, inp, report)

    # Sheet 12: Trailing 12 Month
    _write_trailing_twelve_month(wb, inp, report)

    # Sheet 13: Answer Keys
    _write_answer_keys(wb, inp, report)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
