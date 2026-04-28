"""Parse Damodaran country risk premium spreadsheet (ctryprem.xlsx).

Data on sheet "ERPs by country", header at row 8, data from row 9.
Columns: Country, Region, Moody's rating, Default Spread, Total ERP, Country Risk Premium
"""

from __future__ import annotations

from pathlib import Path

import openpyxl


def parse_country_risk(file_path: str | Path) -> dict[str, dict]:
    """Parse ctryprem.xlsx → dict keyed by country name."""
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb["ERPs by country"]

    # Find header row by looking for "Country" in column A
    header_row = None
    for r in range(1, 15):
        val = ws.cell(r, 1).value
        if val and str(val).strip() == "Country":
            header_row = r
            break
    if header_row is None:
        raise ValueError("Could not find 'Country' header in ERPs by country sheet")

    # Build column map (1-indexed)
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val:
            headers[str(val).strip()] = c

    result: dict[str, dict] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        country = ws.cell(r, headers["Country"]).value
        if not country or not str(country).strip():
            continue
        country = str(country).strip()

        result[country] = {
            "region": _cell_str(ws, r, headers.get("Africa")),  # Column B is region, labeled "Africa" in header
            "moodys_rating": _cell_str(ws, r, headers.get("Moody's rating")),
            "default_spread": _cell_float(ws, r, headers.get("Rating-based Default Spread")),
            "equity_risk_premium": _cell_float(ws, r, headers.get("Total Equity Risk Premium")),
            "country_risk_premium": _cell_float(ws, r, headers.get("Country Risk Premium")),
        }

    # Also extract the mature market ERP from the metadata rows
    mature_erp = None
    for r in range(1, header_row):
        val = ws.cell(r, 1).value
        if val and "mature equity market" in str(val).lower():
            mature_erp = _cell_float(ws, r, 5)  # Column E typically
            break

    if mature_erp is not None:
        result["__mature_market_erp__"] = {"equity_risk_premium": mature_erp}

    return result


def _cell_float(ws, row: int, col: int | None) -> float | None:
    if col is None:
        return None
    val = ws.cell(row, col).value
    if isinstance(val, (int, float)):
        return float(val)
    return None


def _cell_str(ws, row: int, col: int | None) -> str | None:
    if col is None:
        return None
    val = ws.cell(row, col).value
    return str(val).strip() if val else None
