"""
Capital IQ Adapter — unified interface for fetching company financial data.

Strategies:
  1. COM Automation: Drive Excel + CIQ plugin via pywin32 (preferred)
  2. Manual Upload: Parse a pre-filled Excel file uploaded by the user (fallback)

Both strategies produce the same output: RawFinancials + AdjustmentInputs + OptionInputs
mapped to Data Dictionary schemas.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import openpyxl

from engine.data_dictionary import (
    RawFinancials,
    AdjustmentInputs,
    OptionInputs,
)
from .capiq_formula_map import generate_ciq_formulas
from .capiq_excel_automation import ExcelCOMDriver, CIQFetchResult

logger = logging.getLogger(__name__)


class CapIQAdapter:
    """Fetches company financial data via Capital IQ."""

    def __init__(self, years_back: int = 5):
        self.years_back = years_back

    def fetch_via_com(self, ticker: str) -> CapIQResult:
        """
        Fetch data by automating Excel + CIQ plugin.

        Raises RuntimeError if COM automation is not available or fails.
        """
        formulas = generate_ciq_formulas(ticker, years_back=self.years_back)
        driver = ExcelCOMDriver(visible=False, max_wait_sec=90)
        raw_results = driver.fetch(formulas)
        return self._map_results(ticker, raw_results)

    def fetch_via_upload(self, file_path: str | Path) -> CapIQResult:
        """
        Parse a pre-filled Excel file where CIQ formulas have already been resolved.

        The file should have the same layout as the COM automation creates:
        Column A = "variable|period" labels, Column B = resolved values.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        ws = wb.active

        raw_results = []
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=False):
            label_cell, value_cell = row[0], row[1]
            if not label_cell.value:
                continue

            label = str(label_cell.value)
            if "|" not in label:
                continue

            variable, period = label.split("|", 1)
            raw_val = value_cell.value

            result = CIQFetchResult(
                variable=variable.strip(),
                mnemonic="",
                period=period.strip(),
                fiscal_year_offset=_extract_fy_offset(period.strip()),
                value=float(raw_val) if isinstance(raw_val, (int, float)) else None,
                raw_value=raw_val,
                error=None if isinstance(raw_val, (int, float)) else "unparseable",
            )
            raw_results.append(result)

        wb.close()
        return self._map_results("unknown", raw_results)

    def generate_template(self, ticker: str, output_path: str | Path) -> Path:
        """
        Generate an Excel template with CIQ formulas for manual resolution.

        User opens this in Excel with CIQ plugin, formulas resolve, user saves and uploads.
        """
        formulas = generate_ciq_formulas(ticker, years_back=self.years_back)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CIQ_Fetch"

        # Header
        ws.cell(1, 1, "Label (do not edit)")
        ws.cell(1, 2, "CIQ Formula")
        ws.cell(1, 3, "Description")

        for i, f in enumerate(formulas):
            row = i + 2
            ws.cell(row, 1, f"{f['variable']}|{f['period']}")
            # Write the formula as text (CIQ will resolve when opened in Excel)
            ws.cell(row, 2, f["formula"])
            ws.cell(row, 3, f.get("description", ""))

        output_path = Path(output_path)
        wb.save(str(output_path))
        logger.info(f"Generated CIQ template: {output_path} ({len(formulas)} formulas)")
        return output_path

    # Expense-type fields: CIQ returns negative, Damodaran uses positive (absolute value)
    _EXPENSE_FIELDS = {"interest_expense", "capex", "d_a", "operating_lease_expense"}

    def _map_results(self, ticker: str, results: list[CIQFetchResult]) -> CapIQResult:
        """Map raw CIQ results to Data Dictionary Pydantic models."""

        # Group results by fiscal year offset (annual vs quarterly)
        by_year: dict[int, dict[str, float | None]] = {}
        by_quarter: dict[int, dict[str, float | None]] = {}
        current_only: dict[str, float | None] = {}
        period_dates: dict[str, object] = {}  # period_date_annual, period_date_quarterly

        for r in results:
            # Normalize expense-type fields to absolute value (Damodaran convention)
            value = r.value
            if r.variable in self._EXPENSE_FIELDS and value is not None:
                value = abs(value)

            if r.period == "current":
                current_only[r.variable] = value
            elif r.variable.startswith("period_date_"):
                period_dates[r.variable] = r.raw_value
            elif "FQ-" in r.period:
                # Quarterly data
                q_offset = int(r.period.split("FQ-")[1]) if "FQ-" in r.period else 0
                if q_offset not in by_quarter:
                    by_quarter[q_offset] = {}
                by_quarter[q_offset][r.variable] = value
            else:
                yr = r.fiscal_year_offset
                if yr not in by_year:
                    by_year[yr] = {}
                by_year[yr][r.variable] = value

        # Build RawFinancials for each fiscal year
        raw_financials = []
        for yr_offset in sorted(by_year.keys()):
            data = by_year[yr_offset]
            # Merge current-only market data into year 0
            if yr_offset == 0:
                data.update(current_only)

            rf = RawFinancials(
                fiscal_year=datetime.now().year - yr_offset,
                revenues=data.get("revenues", 0.0) or 0.0,
                ebit=data.get("ebit", 0.0) or 0.0,
                ebitda=data.get("ebitda"),
                net_income=data.get("net_income"),
                interest_expense=data.get("interest_expense"),
                capex=data.get("capex"),
                d_a=data.get("d_a"),
                cash_and_marketable_securities=data.get("cash_and_marketable_securities"),
                bv_equity=data.get("bv_equity"),
                bv_debt=data.get("bv_debt"),
                mv_equity=data.get("mv_equity"),
                shares_outstanding=data.get("shares_outstanding"),
                stock_price=data.get("stock_price"),
                cross_holdings=data.get("cross_holdings"),
                minority_interests=data.get("minority_interests"),
            )
            raw_financials.append(rf)

        # Build quarterly RawFinancials (for LTM computation)
        quarterly_financials = []
        for q_offset in sorted(by_quarter.keys()):
            qdata = by_quarter[q_offset]
            qrf = RawFinancials(
                fiscal_year=q_offset,  # 0=most recent quarter, 1=prior, etc.
                revenues=qdata.get("revenues", 0.0) or 0.0,
                ebit=qdata.get("ebit", 0.0) or 0.0,
                ebitda=qdata.get("ebitda"),
                net_income=qdata.get("net_income"),
                interest_expense=qdata.get("interest_expense"),
                capex=qdata.get("capex"),
                d_a=qdata.get("d_a"),
                cash_and_marketable_securities=qdata.get("cash_and_marketable_securities"),
                bv_equity=qdata.get("bv_equity"),
                bv_debt=qdata.get("bv_debt"),
                shares_outstanding=qdata.get("shares_outstanding"),
                cross_holdings=qdata.get("cross_holdings"),
                minority_interests=qdata.get("minority_interests"),
            )
            quarterly_financials.append(qrf)

        # Compute quarters_since_10k from period dates
        quarters_since_10k = 0
        period_date_10k = None
        period_date_10q = None
        pd_annual = period_dates.get("period_date_annual")
        pd_quarterly = period_dates.get("period_date_quarterly")
        if pd_annual is not None:
            period_date_10k = str(pd_annual)
        if pd_quarterly is not None:
            period_date_10q = str(pd_quarterly)
        # Estimate quarters since 10-K: count how many quarterly filings are newer
        if pd_annual and pd_quarterly:
            try:
                from datetime import datetime as dt
                # CIQ dates may come as floats (Excel serial) or strings
                if isinstance(pd_annual, (int, float)) and isinstance(pd_quarterly, (int, float)):
                    diff_days = (pd_quarterly - pd_annual) * 365.25 if pd_quarterly > pd_annual else 0
                    quarters_since_10k = max(1, min(4, round(diff_days / 90)))
            except Exception:
                pass

        # Build AdjustmentInputs from R&D and lease data
        yr0 = by_year.get(0, {})
        r_and_d_past = []
        for yr_offset in range(1, self.years_back + 1):
            yr_data = by_year.get(yr_offset, {})
            rd = yr_data.get("r_and_d_expense")
            r_and_d_past.append(rd or 0.0)

        lease_commitments = []
        for key in ["lease_commitment_yr1", "lease_commitment_yr2", "lease_commitment_yr3",
                     "lease_commitment_yr4", "lease_commitment_yr5", "lease_commitment_beyond"]:
            val = current_only.get(key) or yr0.get(key)
            lease_commitments.append(val or 0.0)

        adj_inputs = AdjustmentInputs(
            amortization_period_n=min(self.years_back, 5),
            r_and_d_expense_current=yr0.get("r_and_d_expense", 0.0) or 0.0,
            r_and_d_expense_past=r_and_d_past,
            operating_lease_expense_current=yr0.get("operating_lease_expense", 0.0) or 0.0,
            operating_lease_commitments=lease_commitments,
            has_r_and_d=(yr0.get("r_and_d_expense") or 0.0) > 0,
            has_operating_leases=(yr0.get("operating_lease_expense") or 0.0) > 0,
        )

        # Build OptionInputs
        option_inputs = OptionInputs(
            number_of_options=current_only.get("options_outstanding", 0.0) or 0.0,
            average_strike_price=current_only.get("options_avg_strike", 0.0) or 0.0,
            average_maturity=current_only.get("options_avg_maturity", 0.0) or 0.0,
            has_options=(current_only.get("options_outstanding") or 0.0) > 0,
        )

        # Collect warnings
        warnings = []
        for r in results:
            if r.error and r.error != "empty":
                warnings.append(f"{r.variable} ({r.period}): {r.error}")

        # Extract reporting currency (comes as a string from CIQ)
        reporting_currency = None
        rc_val = current_only.get("reporting_currency")
        if rc_val is not None:
            reporting_currency = str(rc_val) if not isinstance(rc_val, float) else None

        return CapIQResult(
            ticker=ticker,
            raw_financials=raw_financials,
            adjustment_inputs=adj_inputs,
            option_inputs=option_inputs,
            warnings=warnings,
            reporting_currency=reporting_currency,
            quarterly_financials=quarterly_financials,
            quarters_since_10k=quarters_since_10k,
            period_date_10k=period_date_10k,
            period_date_10q=period_date_10q,
        )


class CapIQResult:
    """Bundled result from a CapIQ data fetch."""

    def __init__(
        self,
        ticker: str,
        raw_financials: list[RawFinancials],
        adjustment_inputs: AdjustmentInputs,
        option_inputs: OptionInputs,
        warnings: list[str] | None = None,
        reporting_currency: str | None = None,
        quarterly_financials: list[RawFinancials] | None = None,
        quarters_since_10k: int = 0,
        period_date_10k: str | None = None,
        period_date_10q: str | None = None,
    ):
        self.ticker = ticker
        self.raw_financials = raw_financials
        self.adjustment_inputs = adjustment_inputs
        self.option_inputs = option_inputs
        self.warnings = warnings or []
        self.reporting_currency = reporting_currency
        self.quarterly_financials = quarterly_financials or []
        self.quarters_since_10k = quarters_since_10k
        self.period_date_10k = period_date_10k
        self.period_date_10q = period_date_10q


def _extract_fy_offset(period: str) -> int:
    """Extract fiscal year offset from period string like 'IQ_FY-2'."""
    if "FY-" in period:
        try:
            return int(period.split("FY-")[1])
        except (ValueError, IndexError):
            pass
    return 0
